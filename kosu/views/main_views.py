from django.shortcuts import render
from django.shortcuts import redirect
from django.contrib import messages
from django.http import HttpResponse
from pathlib import Path
from io import BytesIO
import openpyxl
import datetime
import math
import os
import environ
import urllib.parse
import unicodedata
from ..models import member
from ..models import Business_Time_graph
from ..models import kosu_division
from ..models import team_member
from ..models import inquiry_data
from ..models import administrator_data
from ..forms import loginForm
from ..forms import administrator_data_Form
from ..forms import uploadForm





#--------------------------------------------------------------------------------------------------------





# ログイン画面定義
def login(request):

  # ログイン済みならメイン画面に飛ぶ
  if request.session.get('login_No', None) != None:
    return redirect(to = '/')
  


  # POST時の処理
  if (request.method == 'POST'):

    # POST送信された値を変数に入れる
    find = request.POST['employee_no4']

    # POST送信された値が空の場合の処理
    if find == '':
      # エラーメッセージ出力
      messages.error(request, '従業員番号を入力して下さい。ERROR001')
      # このページをリダイレクト
      return redirect(to = '/login')
    

    # 人員登録データの内、従業員番号がPOST送信された値と等しいレコードのオブジェクトを取得
    data = member.objects.filter(employee_no = find)


    # POST送信された値が人員登録の中にない場合
    if data.count() == 0:
      # エラーメッセージ出力
      messages.error(request, '入力された従業員番号は登録がありません。ERROR002')
      # このページをリダイレクト
      return redirect(to = '/login')
    
    # POST送信された値が人員登録の中にいる場合
    if data.count() >= 1:
      # 従業員番号をセッションに保存する
      request.session['login_No'] = find
      # 使用する工数区分を読み込む(最新のもの)
      def_Ver = kosu_division.objects.order_by("id").last()
      request.session['input_def'] = def_Ver.kosu_name

      # メインページに飛ぶ
      return redirect(to = '/')
   

   
  # HTMLに渡す辞書
  context = {
    'title' : 'ログイン',
    'form' : loginForm(),
    }
  
  # 指定したHTMLに辞書を渡して表示を完成させる
  return render(request, 'kosu/login.html', context)





#--------------------------------------------------------------------------------------------------------





# メイン画面定義
def main(request):

  # 未ログインならログインページに飛ぶ
  if request.session.get('login_No', None) == None:
    return redirect(to = '/login')

  try:
    # ログイン者の情報取得
    data = member.objects.get(employee_no = request.session['login_No'])

  # セッション値から人員情報取得できない場合の処理
  except member.DoesNotExist:
    # セッション削除
    request.session.clear()
    # ログインページに戻る
    return redirect(to = '/login') 

  # 設定データ取得
  default_data = administrator_data.objects.order_by("id").last()

  # 問い合わせ担当者従業員番号がログイン者の従業員番号と一致している場合、ポップアップ表示設定
  if default_data.administrator_employee_no1 == request.session['login_No']:
    pop_up_display = True
  elif default_data.administrator_employee_no2 == request.session['login_No']:
    pop_up_display = True
  elif default_data.administrator_employee_no3 == request.session['login_No']:
    pop_up_display = True

  else:
    pop_up_display = False



  # POST時の処理
  if (request.method == 'POST'):

    # セッションを削除
    request.session.flush()

    # ログインページに飛ぶ
    return redirect(to = '/login')


  
  # HTMLに渡す辞書
  context = {
    'title' : 'MENU',
    'data' : data,
    'pop_up_display' : pop_up_display,
    'pop_up' : default_data,
    }
  


  # 指定したHTMLに辞書を渡して表示を完成させる
  return render(request, 'kosu/main.html', context)





#--------------------------------------------------------------------------------------------------------





# 工数メイン画面定義
def kosu_main(request):

  # 未ログインならログインページに飛ぶ
  if request.session.get('login_No', None) == None:

    return redirect(to = '/login')


  try:
    # ログイン者の情報取得
    data = member.objects.get(employee_no = request.session['login_No'])

  # セッション値から人員情報取得できない場合の処理
  except member.DoesNotExist:
    # セッション削除
    request.session.clear()
    # ログインページに戻る
    return redirect(to = '/login') 



  # HTMLに渡す辞書
  context = {
    'title' : '工数MENU',
    'data' : data,
    }
  


  # 指定したHTMLに辞書を渡して表示を完成させる
  return render(request, 'kosu/kosu_main.html', context)





#--------------------------------------------------------------------------------------------------------





# 工数区分定義メイン画面定義
def def_main(request):

  # 未ログインならログインページに飛ぶ
  if request.session.get('login_No', None) == None:

    return redirect(to = '/login')

  try:
    # ログイン者の情報取得
    data = member.objects.get(employee_no = request.session['login_No'])

  # セッション値から人員情報取得できない場合の処理
  except member.DoesNotExist:
    # セッション削除
    request.session.clear()
    # ログインページに戻る
    return redirect(to = '/login') 



  # HTMLに渡す辞書
  context = {
    'title' : '工数区分定義MENU',
    'data' : data,
    }
  


  # 指定したHTMLに辞書を渡して表示を完成させる
  return render(request, 'kosu/def_main.html', context)





#--------------------------------------------------------------------------------------------------------





# 人員メイン画面定義
def member_main(request):

  # 未ログインならログインページに飛ぶ
  if request.session.get('login_No', None) == None:

    return redirect(to = '/login')

  try:
    # ログイン者の情報取得
    data = member.objects.get(employee_no = request.session['login_No'])

  # セッション値から人員情報取得できない場合の処理
  except member.DoesNotExist:
    # セッション削除
    request.session.clear()
    # ログインページに戻る
    return redirect(to = '/login') 

  # ログイン者に権限がなければメインページに戻る
  if data.authority == False:

    return redirect(to = '/')
  


  # HTMLに渡す辞書
  context = {
    'title' : '人員MENU',
    'data' : data,
    }
  


  # 指定したHTMLに辞書を渡して表示を完成させる
  return render(request, 'kosu/member_main.html', context)





#--------------------------------------------------------------------------------------------------------





# 班員メイン画面定義
def team_main(request):

  # 未ログインならログインページに飛ぶ
  if request.session.get('login_No', None) == None:
    return redirect(to = '/login')


  try:
    # ログイン者の情報取得
    data = member.objects.get(employee_no = request.session['login_No'])

  # セッション値から人員情報取得できない場合の処理
  except member.DoesNotExist:
    # セッション削除
    request.session.clear()
    # ログインページに戻る
    return redirect(to = '/login') 

  # ログイン者に権限がなければメインページに戻る
  if data.authority == False:

    return redirect(to = '/')


  # 今日の日時を変数に格納
  today = datetime.date.today()

  # フォロー表示変数定義
  follow_display = False

  # 日付リスト初期状態定義
  day_list = []
  # 日付リスト作成ループ
  for d in range(1, 8):
    # 日付リストに今日から1週間前までの日付を入れる
    day_list.append(today - datetime.timedelta(days=d))

  # ログイン者の班員登録あるか確認
  team_filter = team_member.objects.filter(employee_no5 = request.session['login_No'])

  # ログイン者の班員登録がある場合の処理
  if team_filter.count() != 0:
    # ログイン者の班員情報取得
    team_get = team_member.objects.get(employee_no5 = request.session['login_No'])

    # フォロー表示
    follow_display = team_get.follow

    # 班員フォロー用リスト定義
    team_list = []
    # 班員フォロー用リスト作成するループ
    for t in range(1, 16):
      # ログイン者の班員情報がある場合の処理
      if eval('team_get.member{}'.format(t)) != '':
        # 班員が人員情報にいるか確認
        member_name_filter = member.objects.filter(employee_no =  eval('team_get.member{}'.format(t)))

        # 人員登録がある場合の処理
        if member_name_filter.count() != 0:
          # 班員の人員情報取得
          member_name_get = member.objects.get(employee_no =  eval('team_get.member{}'.format(t)))

          # 班員の過去1週間の工数データが正しく入力されているか確認するループ
          for ind, dd in enumerate(day_list):
            # 班員の工数データが該当日にあるか確認
            kosu_filter = Business_Time_graph.objects.filter(employee_no3 = eval('team_get.member{}'.format(t)), \
                                                             work_day2 = dd)
            
            # 班員の工数データが該当日にあった場合の処理
            if kosu_filter.count() != 0:
              # 班員の該当日の工数データ取得
              kosu_get = Business_Time_graph.objects.get(employee_no3 = eval('team_get.member{}'.format(t)), \
                                                         work_day2 = dd)
              
              # 昨日の工数データが3直以外で整合性NGの場合の処理
              if kosu_get.tyoku2 != '3' and ind == 0 and kosu_get.judgement != True:
                # 班員フォーローリストにコメント追加
                team_list.append('{}氏の工数未入力があります。'.format(member_name_get.name))
                break

              # 2日以前の工数データが整合性NGの場合の処理
              if kosu_get.judgement != True and ind != 0:
                # 班員フォーローリストにコメント追加
                team_list.append('{}氏の工数未入力があります。'.format(member_name_get.name))
                break

            # 班員の工数データが該当日にない場合の処理
            else:
              # 班員フォーローリストにコメント追加
              team_list.append('{}氏の工数未入力があります。'.format(member_name_get.name))
              break

  # ログイン者の班員登録がない場合の処理
  else:
    # 班員情報に空を入れる
    team_list = []



  # HTMLに渡す辞書
  context = {
    'title' : '班員MENU',
    'data' : data,
    'team' : team_list,
    'follow_display' : follow_display,
    }



  # 指定したHTMLに辞書を渡して表示を完成させる
  return render(request, 'kosu/team_main.html', context)





#--------------------------------------------------------------------------------------------------------





# 問い合わせメイン画面定義
def inquiry_main(request):

  # 未ログインならログインページに飛ぶ
  if request.session.get('login_No', None) == None:

    return redirect(to = '/login')


  try:
    # ログイン者の情報取得
    data = member.objects.get(employee_no = request.session['login_No'])

  # セッション値から人員情報取得できない場合の処理
  except member.DoesNotExist:
    # セッション削除
    request.session.clear()
    # ログインページに戻る
    return redirect(to = '/login') 



  # HTMLに渡す辞書
  context = {
    'title' : '問い合わせMENU',
    'data' : data,
    }
  


  # 指定したHTMLに辞書を渡して表示を完成させる
  return render(request, 'kosu/inquiry_main.html', context)





#--------------------------------------------------------------------------------------------------------





# 管理者画面定義
def administrator_menu(request):

  # 未ログインならログインページに飛ぶ
  if request.session.get('login_No', None) == None:
    return redirect(to = '/login')

  try:
    # ログイン者の情報取得
    data = member.objects.get(employee_no = request.session['login_No'])

  # セッション値から人員情報取得できない場合の処理
  except member.DoesNotExist:
    # セッション削除
    request.session.clear()
    # ログインページに戻る
    return redirect(to = '/login') 


  # ログイン者が管理者でなければメインページに戻る
  if data.administrator == False:
    return redirect(to = '/')


  # 設定データ取得
  default_data = administrator_data.objects.order_by("id").last()

  # 設定データの最新のレコードのID取得
  record_id = default_data.id


  # フォーム初期値
  form_default = {
    'menu_row' : default_data.menu_row,
    'administrator_employee_no1' : default_data.administrator_employee_no1,
    'administrator_employee_no2' : default_data.administrator_employee_no2,
    'administrator_employee_no3' : default_data.administrator_employee_no3,
    }
  
  # パス指定フォームに初期値を入れて定義
  form = administrator_data_Form(form_default)

  # ロードファイル指定フォーム定義
  load_form = uploadForm()
    

  # 設定更新時の処理
  if 'registration' in request.POST:
    # 一覧表示項目数に文字が入っている場合の処理
    if not request.POST['menu_row'].isdigit():
      # エラーメッセージ出力
      messages.error(request, '一覧表示項目数は数字で入力して下さい。ERROR056')
      # このページをリダイレクト
      return redirect(to = '/administrator')

    # 問い合わせ担当者従業員番号が空でない場合の処理
    if request.POST['administrator_employee_no1'] != '':
      # 問い合わせ担当者従業員番号に文字が入っている場合の処理
      if not request.POST['administrator_employee_no1'].isdigit():
        # エラーメッセージ出力
        messages.error(request, '問い合わせ担当者従業員番号は数字で入力して下さい。ERROR077')
        # このページをリダイレクト
        return redirect(to = '/administrator')

    # 問い合わせ担当者従業員番号が空でない場合の処理
    if request.POST['administrator_employee_no2'] != '':
      # 問い合わせ担当者従業員番号に文字が入っている場合の処理
      if not request.POST['administrator_employee_no2'].isdigit():
        # エラーメッセージ出力
        messages.error(request, '問い合わせ担当者従業員番号は数字で入力して下さい。ERROR078')
        # このページをリダイレクト
        return redirect(to = '/administrator')

    # 問い合わせ担当者従業員番号が空でない場合の処理
    if request.POST['administrator_employee_no3'] != '':
      # 問い合わせ担当者従業員番号に文字が入っている場合の処理
      if not request.POST['administrator_employee_no3'].isdigit():
        # エラーメッセージ出力
        messages.error(request, '問い合わせ担当者従業員番号は数字で入力して下さい。ERROR079')
        # このページをリダイレクト
        return redirect(to = '/administrator')

    # 一覧表示項目数が自然数でない場合の処理
    if math.floor(float(request.POST['menu_row'])) != float(request.POST['menu_row']) or \
      float(request.POST['menu_row']) <= 0:
      # エラーメッセージ出力
      messages.error(request, '一覧表示項目数は自然数で入力して下さい。ERROR029')
      # このページをリダイレクト
      return redirect(to = '/administrator')
    
    # 問い合わせ担当者従業員番号が空でない場合の処理
    if request.POST['administrator_employee_no1'] != '':
      # 問い合わせ担当者従業員番号が自然数でない場合の処理
      if math.floor(float(request.POST['administrator_employee_no1'])) != \
        float(request.POST['administrator_employee_no1']) or \
        float(request.POST['administrator_employee_no1']) <= 0:
        # エラーメッセージ出力
        messages.error(request, '問い合わせ担当者従業員番号は自然数で入力して下さい。ERROR033')
        # このページをリダイレクト
        return redirect(to = '/administrator')

    # 問い合わせ担当者従業員番号が空でない場合の処理
    if request.POST['administrator_employee_no2'] != '':
      # 問い合わせ担当者従業員番号が自然数でない場合の処理
      if math.floor(float(request.POST['administrator_employee_no2'])) != \
        float(request.POST['administrator_employee_no2']) or \
        float(request.POST['administrator_employee_no2']) <= 0:
        # エラーメッセージ出力
        messages.error(request, '問い合わせ担当者従業員番号は自然数で入力して下さい。ERROR034')
        # このページをリダイレクト
        return redirect(to = '/administrator')

    # 問い合わせ担当者従業員番号が空でない場合の処理
    if request.POST['administrator_employee_no3'] != '':
      # 問い合わせ担当者従業員番号が自然数でない場合の処理
      if math.floor(float(request.POST['administrator_employee_no3'])) != \
        float(request.POST['administrator_employee_no3']) or \
        float(request.POST['administrator_employee_no3']) <= 0:
        # エラーメッセージ出力
        messages.error(request, '問い合わせ担当者従業員番号は自然数で入力して下さい。ERROR035')
        # このページをリダイレクト
        return redirect(to = '/administrator')

    # 一覧表示項目数が半角でない場合の処理
    if has_non_halfwidth_characters(request.POST['menu_row']):
      # エラーメッセージ出力
      messages.error(request, '一覧表示項目数は半角で入力して下さい。ERROR053')
      # このページをリダイレクト
      return redirect(to = '/administrator')

    # 問い合わせ担当者従業員番号が半角でない場合の処理
    if has_non_halfwidth_characters(request.POST['administrator_employee_no1']):
      # エラーメッセージ出力
      messages.error(request, '問い合わせ担当者従業員番号は半角で入力して下さい。ERROR036')
      # このページをリダイレクト
      return redirect(to = '/administrator')

    # 問い合わせ担当者従業員番号が半角でない場合の処理
    if has_non_halfwidth_characters(request.POST['administrator_employee_no2']):
      # エラーメッセージ出力
      messages.error(request, '問い合わせ担当者従業員番号は半角で入力して下さい。ERROR037')
      # このページをリダイレクト
      return redirect(to = '/administrator')

    # 問い合わせ担当者従業員番号が半角でない場合の処理
    if has_non_halfwidth_characters(request.POST['administrator_employee_no3']):
      # エラーメッセージ出力
      messages.error(request, '問い合わせ担当者従業員番号は半角で入力して下さい。ERROR038')
      # このページをリダイレクト
      return redirect(to = '/administrator')

    # 問い合わせ担当者従業員番号が空でない場合の処理
    if request.POST['administrator_employee_no1'] != '':
      # 送信された問い合わせ担当者従業員番号の人員データあるか確認
      member_obj_filter = member.objects.filter(employee_no = request.POST['administrator_employee_no1'])
      # 人員データ無い場合の処理
      if member_obj_filter.count() == 0:
        # エラーメッセージ出力
        messages.error(request, '入力された問い合わせ担当者従業員番号は登録されていません。ERROR055')
        # このページをリダイレクト
        return redirect(to = '/administrator')

    # 問い合わせ担当者従業員番号が空でない場合の処理
    if request.POST['administrator_employee_no2'] != '':
      # 送信された問い合わせ担当者従業員番号の人員データあるか確認
      member_obj_filter = member.objects.filter(employee_no = request.POST['administrator_employee_no2'])
      # 人員データ無い場合の処理
      if member_obj_filter.count() == 0:
        # エラーメッセージ出力
        messages.error(request, '入力された問い合わせ担当者従業員番号は登録されていません。ERROR080')
        # このページをリダイレクト
        return redirect(to = '/administrator')

    # 問い合わせ担当者従業員番号が空でない場合の処理
    if request.POST['administrator_employee_no3'] != '':
      # 送信された問い合わせ担当者従業員番号の人員データあるか確認
      member_obj_filter = member.objects.filter(employee_no = request.POST['administrator_employee_no3'])
      # 人員データ無い場合の処理
      if member_obj_filter.count() == 0:
        # エラーメッセージ出力
        messages.error(request, '入力された問い合わせ担当者従業員番号は登録されていません。ERROR081')
        # このページをリダイレクト
        return redirect(to = '/administrator')

    # レコードにPOST送信された値を上書きする
    administrator_data.objects.update_or_create(id = record_id, \
        defaults = {'menu_row' : request.POST['menu_row'], 
                    'administrator_employee_no1' : request.POST['administrator_employee_no1'],
                    'administrator_employee_no2' : request.POST['administrator_employee_no2'],
                    'administrator_employee_no3' : request.POST['administrator_employee_no3']})
    
    # フォームにPOST値を入れて定義
    form = administrator_data_Form(request.POST)



  # 工数情報バックアップ処理
  if 'kosu_backup' in request.POST:
    # 日付指定空の場合の処理
    if request.POST['data_day'] == '':
      # エラーメッセージ出力
      messages.error(request, 'バックアップする日付を指定してください。ERROR022')
      # このページをリダイレクト
      return redirect(to = '/administrator')


    # 新しいExcelブック作成
    wb = openpyxl.Workbook()
    # 書き込みシート選択
    ws = wb.active
    # 工数データ取得
    kosu_data = Business_Time_graph.objects.filter(work_day2__lte = request.POST['data_day'])

    # Excelに書き込み(項目名)
    headers = [
      '従業員番号', 
      '氏名', 
      '工数区分定義Ver', 
      '就業日', 
      '直', 
      '作業内容',
      '作業詳細', 
      '残業時間', 
      '昼休憩時間', 
      '残業休憩時間1', 
      '残業休憩時間2',
      '残業休憩時間3', 
      '就業形態', 
      '工数入力OK_NG'
      ]
    ws.append(headers)

    # Excelに書き込み(データ)
    for item in kosu_data:

      row = [
        item.employee_no3, 
        str(item.name), 
        item.def_ver2, 
        item.work_day2,
        item.tyoku2, 
        item.time_work, 
        item.detail_work, 
        item.over_time,
        item.breaktime, 
        item.breaktime_over1, 
        item.breaktime_over2,
        item.breaktime_over3, 
        item.work_time, 
        item.judgement
        ]
      ws.append(row)

    # メモリ上にExcelファイルを作成し、BytesIOオブジェクトに保存
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # ファイル名を設定
    filename = f'工数データバックアップ_{request.POST["data_day"]}.xlsx'

    # URLエンコーディングされたファイル名を生成
    quoted_filename = urllib.parse.quote(filename)
    

    # HttpResponseを作成してファイルをダウンロードさせる
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # Content-Dispositionヘッダーを設定
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{quoted_filename}'
    
    return response



  # 工数データ読み込み
  if 'kosu_load' in request.POST:
    # 工数データファイルが未選択時の処理
    if 'kosu_file' not in request.FILES:
      # エラーメッセージ出力
      messages.error(request, '工数データファイルが選択されていません。ERROR040')
      # このページをリダイレクト
      return redirect(to = '/administrator')

    # POSTされたファイルパスを変数に入れる
    file_path = request.FILES['kosu_file']


    # 一時的なファイルをサーバー上に作成
    with open('kosu_file_path.xlsx', 'wb+') as destination:
      # アップロードしたファイルを一時ファイルに書き込み
      for chunk in file_path.chunks():
        destination.write(chunk)

    # 指定Excelを開く
    wb = openpyxl.load_workbook('kosu_file_path.xlsx')
    # 書き込みシート選択
    ws = wb.worksheets[0]

    # 読み込むファイルが正しいファイルでない場合の処理
    if ws.cell(1, 1).value != '従業員番号' or ws.cell(1, 2).value != '氏名' or \
      ws.cell(1, 3).value != '工数区分定義Ver' or ws.cell(1, 4).value != '就業日' or \
      ws.cell(1, 5).value != '直' or ws.cell(1, 6).value != '作業内容' or \
      ws.cell(1, 7).value != '作業詳細' or ws.cell(1, 8).value != '残業時間' or \
      ws.cell(1, 9).value != '昼休憩時間' or ws.cell(1, 10).value != '残業休憩時間1' or \
      ws.cell(1, 11).value != '残業休憩時間2' or ws.cell(1, 12).value != '残業休憩時間3' or \
      ws.cell(1, 13).value != '就業形態' or ws.cell(1, 14).value != '工数入力OK_NG':

      # エラーメッセージ出力
      messages.error(request, 'ロードしようとしたファイルは工数データバックアップではありません。ERROR048')
      # このページをリダイレクト
      return redirect(to = '/administrator')

    # レコード数取得
    data_num = ws.max_row

    # Excelからデータを読み込こむループ
    for i in range(1, data_num):
      # 読み込み予定データと同一の日のデータが存在するか確認
      kosu_data_filter = Business_Time_graph.objects.filter(employee_no3 = ws.cell(row = i + 1, column = 1).value, \
                                                           work_day2 = ws.cell(row = i + 1, column = 4).value)
      # 読み込み予定データと同一の日のデータが存在する場合の処理
      if kosu_data_filter.count() != 0:
        # 読み込み予定データと同一の日のデータを取得
        kosu_data_get = Business_Time_graph.objects.get(employee_no3 = ws.cell(row = i + 1, column = 1).value, \
                                                       work_day2 = ws.cell(row = i + 1, column = 4).value)
        # 読み込み予定データと同一の日のデータを削除
        kosu_data_get.delete()

      # 人員データインスタンス取得
      member_instance = member.objects.get(employee_no = ws.cell(row = i + 1, column = 1).value)

      # Excelからデータを読み込こみ
      new_data = Business_Time_graph(employee_no3 = ws.cell(row = i + 1, column = 1).value, \
                                      name = member_instance, \
                                      def_ver2 = ws.cell(row = i + 1, column = 3).value, \
                                      work_day2 = ws.cell(row = i + 1, column = 4).value, \
                                      tyoku2 = ws.cell(row = i + 1, column = 5).value, \
                                      time_work = ws.cell(row = i + 1, column = 6).value, \
                                      detail_work = ws.cell(row = i + 1, column = 7).value, \
                                      over_time = ws.cell(row = i + 1, column = 8).value, \
                                      breaktime = ws.cell(row = i + 1, column = 9).value, \
                                      breaktime_over1 = ws.cell(row = i + 1, column = 10).value, \
                                      breaktime_over2 = ws.cell(row = i + 1, column = 11).value, \
                                      breaktime_over3 = ws.cell(row = i + 1, column = 12).value, \
                                      work_time = ws.cell(row = i + 1, column = 13).value, \
                                      judgement = ws.cell(row = i + 1, column = 14).value) 

      # レコードセーブ
      new_data.save()

    # 一時ファイル削除
    os.remove('kosu_file_path.xlsx')



  # 工数データ削除
  if 'kosu_delete' in request.POST:
    # 日付指定空の場合の処理
    if request.POST['data_day'] == '':
      # エラーメッセージ出力
      messages.error(request, '削除する日付を指定してください。ERROR023')
      # このページをリダイレクト
      return redirect(to = '/administrator')


    # 工数データ取得
    kosu_obj = Business_Time_graph.objects.filter(work_day2__lte = request.POST['data_day'])
    # 取得した工数データを削除
    kosu_obj.delete()



  # 人員情報バックアップ処理
  if 'member_backup' in request.POST:
    # 今日の日付取得
    today = datetime.date.today().strftime('%Y%m%d')
    # 新しいExcelブック作成
    wb = openpyxl.Workbook()
    # 書き込みシート選択
    ws = wb.active
    # 人員データ取得
    member_data = member.objects.all()

    # Excelに書き込み(項目名)
    headers = [
        '従業員番号', 
        '氏名', 
        'ショップ', 
        '権限', 
        '管理者', 
        '1直昼休憩時間',
        '1直残業休憩時間1', 
        '1直残業休憩時間2', 
        '1直残業休憩時間3', 
        '2直昼休憩時間', 
        '2直残業休憩時間1',
        '2直残業休憩時間2', 
        '2直残業休憩時間3', 
        '3直昼休憩時間', 
        '3直残業休憩時間1', 
        '3直残業休憩時間2', 
        '3直残業休憩時間3', 
        '常昼昼休憩時間', 
        '常昼残業休憩時間1', 
        '常昼残業休憩時間2', 
        '常昼残業休憩時間3',
        ]
    ws.append(headers)

    # Excelに書き込み(データ)
    for item in member_data:
      row = [
        item.employee_no, 
        item.name, 
        item.shop, 
        item.authority,
        item.administrator, 
        item.break_time1, 
        item.break_time1_over1, 
        item.break_time1_over2, 
        item.break_time1_over3, 
        item.break_time2, 
        item.break_time2_over1, 
        item.break_time2_over2, 
        item.break_time2_over3, 
        item.break_time3, 
        item.break_time3_over1, 
        item.break_time3_over2, 
        item.break_time3_over3, 
        item.break_time4, 
        item.break_time4_over1, 
        item.break_time4_over2, 
        item.break_time4_over3,
        ]
      ws.append(row)


    # メモリ上にExcelファイルを作成
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # ファイル名を設定
    filename = f'人員データバックアップ_{today}.xlsx'

    # URLエンコーディングされたファイル名を生成
    quoted_filename = urllib.parse.quote(filename)
    
    # HttpResponseを作成してファイルをダウンロードさせる
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # Content-Dispositionヘッダーを設定
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{quoted_filename}'
    
    return response



  # 人員情報読み込み
  if 'member_load' in request.POST:
    # 人員データファイルが未選択時の処理
    if 'member_file' not in request.FILES:
      # エラーメッセージ出力
      messages.error(request, '人員ファイルが選択されていません。ERROR044')
      # このページをリダイレクト
      return redirect(to = '/administrator')


    # POSTされたファイルパスを変数に入れる
    file_path = request.FILES['member_file']


    # 一時的なファイルをサーバー上に作成
    with open('member_file_path.xlsx', 'wb+') as destination:
      # アップロードしたファイルを一時ファイルに書き込み
      for chunk in file_path.chunks():
        destination.write(chunk)

    # 指定Excelを開く
    wb = openpyxl.load_workbook('member_file_path.xlsx')
    # 書き込みシート選択
    ws = wb.worksheets[0]


    # 読み込むファイルが正しいファイルでない場合の処理
    if ws.cell(1, 1).value != '従業員番号' or ws.cell(1, 2).value != '氏名' or \
      ws.cell(1, 3).value != 'ショップ' or ws.cell(1, 4).value != '権限' or \
      ws.cell(1, 5).value != '管理者' or ws.cell(1, 6).value != '1直昼休憩時間' or \
      ws.cell(1, 7).value != '1直残業休憩時間1' or ws.cell(1, 8).value != '1直残業休憩時間2' or \
      ws.cell(1, 9).value != '1直残業休憩時間3' or ws.cell(1, 10).value != '2直昼休憩時間' or \
      ws.cell(1, 11).value != '2直残業休憩時間1' or ws.cell(1, 12).value != '2直残業休憩時間2' or \
      ws.cell(1, 13).value != '2直残業休憩時間3' or ws.cell(1, 14).value != '3直昼休憩時間' or \
      ws.cell(1, 15).value != '3直残業休憩時間1' or ws.cell(1, 16).value != '3直残業休憩時間2' or \
      ws.cell(1, 17).value != '3直残業休憩時間3' or ws.cell(1, 18).value != '常昼昼休憩時間' or \
      ws.cell(1, 19).value != '常昼残業休憩時間1' or ws.cell(1, 20).value != '常昼残業休憩時間2' or \
      ws.cell(1, 21).value != '常昼残業休憩時間3':
      # エラーメッセージ出力
      messages.error(request, 'ロードしようとしたファイルは人員情報バックアップではありません。ERROR049')
      # このページをリダイレクト
      return redirect(to = '/administrator')

    # レコード数取得
    data_num = ws.max_row


    # Excelからデータを読み込むループ
    for i in range(1, data_num):
      # 読み込み予定データと同一の従業員番号のデータが存在するか確認
      member_data_filter = member.objects.filter(employee_no = ws.cell(row = i + 1, column = 1).value)

      # 上書きチェックONの場合の処理
      if ('overwrite_check' in request.POST):
        # 読み込み予定データと同一の従業員番号のデータが存在する場合の処理
        if member_data_filter.count() != 0:
          # 読み込み予定データと同一の従業員番号のデータを取得
          member_data_get = member.objects.get(employee_no = ws.cell(row = i + 1, column = 1).value)
          # 読み込み予定データと同一の従業員番号のデータを削除
          member_data_get.delete()

        # Excelからデータを読み込み
        new_data = member(employee_no = ws.cell(row = i + 1, column = 1).value, \
                          name = ws.cell(row = i + 1, column = 2).value, \
                          shop = ws.cell(row = i + 1, column = 3).value, \
                          authority = ws.cell(row = i + 1, column = 4).value, \
                          administrator = ws.cell(row = i + 1, column = 5).value, \
                          break_time1 = ws.cell(row = i + 1, column = 6).value, \
                          break_time1_over1 = ws.cell(row = i + 1, column = 7).value, \
                          break_time1_over2 = ws.cell(row = i + 1, column = 8).value, \
                          break_time1_over3 = ws.cell(row = i + 1, column = 9).value, \
                          break_time2 = ws.cell(row = i + 1, column = 10).value, \
                          break_time2_over1 = ws.cell(row = i + 1, column = 11).value, \
                          break_time2_over2 = ws.cell(row = i + 1, column = 12).value, \
                          break_time2_over3 = ws.cell(row = i + 1, column = 13).value, \
                          break_time3 = ws.cell(row = i + 1, column = 14).value, \
                          break_time3_over1 = ws.cell(row = i + 1, column = 15).value, \
                          break_time3_over2 = ws.cell(row = i + 1, column = 16).value, \
                          break_time3_over3 = ws.cell(row = i + 1, column = 17).value, \
                          break_time4 = ws.cell(row = i + 1, column = 18).value, \
                          break_time4_over1 = ws.cell(row = i + 1, column = 19).value, \
                          break_time4_over2 = ws.cell(row = i + 1, column = 20).value, \
                          break_time4_over3 = ws.cell(row = i + 1, column = 21).value)
        new_data.save()

      # 上書きチェックOFFの場合の処理
      else:
        # 読み込み予定データと同一の従業員番号のデータが存在しない場合の処理
        if member_data_filter.count() == 0:
          # Excelからデータを読み込み
          new_data = member(employee_no = ws.cell(row = i + 1, column = 1).value, \
                            name = ws.cell(row = i + 1, column = 2).value, \
                            shop = ws.cell(row = i + 1, column = 3).value, \
                            authority = ws.cell(row = i + 1, column = 4).value, \
                            administrator = ws.cell(row = i + 1, column = 5).value, \
                            break_time1 = ws.cell(row = i + 1, column = 6).value, \
                            break_time1_over1 = ws.cell(row = i + 1, column = 7).value, \
                            break_time1_over2 = ws.cell(row = i + 1, column = 8).value, \
                            break_time1_over3 = ws.cell(row = i + 1, column = 9).value, \
                            break_time2 = ws.cell(row = i + 1, column = 10).value, \
                            break_time2_over1 = ws.cell(row = i + 1, column = 11).value, \
                            break_time2_over2 = ws.cell(row = i + 1, column = 12).value, \
                            break_time2_over3 = ws.cell(row = i + 1, column = 13).value, \
                            break_time3 = ws.cell(row = i + 1, column = 14).value, \
                            break_time3_over1 = ws.cell(row = i + 1, column = 15).value, \
                            break_time3_over2 = ws.cell(row = i + 1, column = 16).value, \
                            break_time3_over3 = ws.cell(row = i + 1, column = 17).value, \
                            break_time4 = ws.cell(row = i + 1, column = 18).value, \
                            break_time4_over1 = ws.cell(row = i + 1, column = 19).value, \
                            break_time4_over2 = ws.cell(row = i + 1, column = 20).value, \
                            break_time4_over3 = ws.cell(row = i + 1, column = 21).value)
          new_data.save()

    # 一時ファイル削除
    os.remove('member_file_path.xlsx')



  # 班員情報バックアップ処理
  if 'team_backup' in request.POST:
    # 今日の日付取得
    today = datetime.date.today().strftime('%Y%m%d')
    # 新しいExcelブック作成
    wb = openpyxl.Workbook()
    # 書き込みシート選択
    ws = wb.active
    # 班員データ取得
    team_data = team_member.objects.all()

    # Excelに書き込み(項目名)
    headers = [
        '従業員番号', 
        '班員1', 
        '班員2', 
        '班員3', 
        '班員4', 
        '班員5',
        '班員6', 
        '班員7', 
        '班員8', 
        '班員9', 
        '班員10',
        ]
    ws.append(headers)

    # Excelに書き込み(データ)
    for item in team_data:
      row = [
        item.employee_no5, 
        item.member1, 
        item.member2, 
        item.member3, 
        item.member4, 
        item.member5, 
        item.member6, 
        item.member7, 
        item.member8, 
        item.member9, 
        item.member10
        ]
      ws.append(row)

    # メモリ上にExcelファイルを作成
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    # ファイル名を設定
    filename = f'班員データバックアップ_{today}.xlsx'
    # URLエンコーディングされたファイル名を生成
    quoted_filename = urllib.parse.quote(filename)
    
    # HttpResponseを作成してファイルをダウンロードさせる
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # Content-Dispositionヘッダーを設定
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{quoted_filename}'
    
    return response



  # 班員情報読み込み
  if 'team_load' in request.POST:
    # 工数データファイルが未選択時の処理
    if 'team_file' not in request.FILES:
      # エラーメッセージ出力
      messages.error(request, '班員ファイルが選択されていません。ERROR045')
      # このページをリダイレクト
      return redirect(to = '/administrator')


    # POSTされたファイルパスを変数に入れる
    file_path = request.FILES['team_file']

    # 一時的なファイルをサーバー上に作成
    with open('team_file_path.xlsx', 'wb+') as destination:

      # アップロードしたファイルを一時ファイルに書き込み
      for chunk in file_path.chunks():
        destination.write(chunk)

    # 指定Excelを開く
    wb = openpyxl.load_workbook('team_file_path.xlsx')
    # 書き込みシート選択
    ws = wb.worksheets[0]

    # 読み込むファイルが正しいファイルでない場合の処理
    if ws.cell(1, 1).value != '従業員番号' or ws.cell(1, 2).value != '班員1' or \
      ws.cell(1, 3).value != '班員2' or ws.cell(1, 4).value != '班員3' or \
      ws.cell(1, 5).value != '班員4' or ws.cell(1, 6).value != '班員5' or \
      ws.cell(1, 7).value != '班員6' or ws.cell(1, 8).value != '班員7' or \
      ws.cell(1, 9).value != '班員8' or ws.cell(1, 10).value != '班員9' or \
      ws.cell(1, 11).value != '班員10':
      # エラーメッセージ出力
      messages.error(request, 'ロードしようとしたファイルは班員情報バックアップではありません。ERROR050')
      # このページをリダイレクト
      return redirect(to = '/administrator')

    # レコード数取得
    data_num = ws.max_row

    # Excelからデータを読み込むループ
    for i in range(1, data_num):
      # 読み込み予定データと同一の従業員番号のデータが存在するか確認
      team_data_filter = team_member.objects.filter(employee_no5 = ws.cell(row = i + 1, column = 1).value)

      # 読み込み予定データと同一の従業員番号のデータが存在する場合の処理
      if team_data_filter.count() != 0:
        # 読み込み予定データと同一の従業員番号のデータを取得
        team_data_get = team_member.objects.get(employee_no5 = ws.cell(row = i + 1, column = 1).value)
        # 読み込み予定データと同一の従業員番号のデータを削除
        team_data_get.delete()

      # Excelからデータ読み込み
      new_data = team_member(employee_no5 = ws.cell(row = i + 1, column = 1).value, \
                             member1 = ws.cell(row = i + 1, column = 2).value, \
                             member2 = ws.cell(row = i + 1, column = 3).value, \
                             member3 = ws.cell(row = i + 1, column = 4).value, \
                             member4 = ws.cell(row = i + 1, column = 5).value, \
                             member5 = ws.cell(row = i + 1, column = 6).value, \
                             member6 = ws.cell(row = i + 1, column = 7).value, \
                             member7 = ws.cell(row = i + 1, column = 8).value, \
                             member8 = ws.cell(row = i + 1, column = 9).value, \
                             member9 = ws.cell(row = i + 1, column = 10).value, \
                             member10 = ws.cell(row = i + 1, column = 11).value) 

      new_data.save()
 
    # 一時ファイル削除
    os.remove('team_file_path.xlsx')



  # 工数区分定義バックアップ処理
  if 'def_backup' in request.POST:
    # 今日の日付取得
    today = datetime.date.today().strftime('%Y%m%d')
    # 新しいExcelブック作成
    wb = openpyxl.Workbook()
    # 書き込みシート選択
    ws = wb.active

    # メンバーデータ取得
    def_data = kosu_division.objects.all()


    # Excelに書き込み(項目名)
    headers = [
        '工数区分定義Ver名', '工数区分名1', '定義1', '作業内容1', '工数区分名2', '定義2', '作業内容2',
        '工数区分名3', '定義3', '作業内容3', '工数区分名4', '定義4', '作業内容4',
        '工数区分名5', '定義5', '作業内容5', '工数区分名6', '定義6', '作業内容6',
        '工数区分名7', '定義7', '作業内容7', '工数区分名8', '定義8', '作業内容8',
        '工数区分名9', '定義9', '作業内容9', '工数区分名10', '定義10', '作業内容10',
        '工数区分名11', '定義11', '作業内容11', '工数区分名12', '定義12', '作業内容12',
        '工数区分名13', '定義13', '作業内容13', '工数区分名14', '定義14', '作業内容14',
        '工数区分名15', '定義15', '作業内容15', '工数区分名16', '定義16', '作業内容16',
        '工数区分名17', '定義17', '作業内容17', '工数区分名18', '定義18', '作業内容18',
        '工数区分名19', '定義19', '作業内容19', '工数区分名20', '定義20', '作業内容20',
        '工数区分名21', '定義21', '作業内容21', '工数区分名22', '定義22', '作業内容22',
        '工数区分名23', '定義23', '作業内容23', '工数区分名24', '定義24', '作業内容24',
        '工数区分名25', '定義25', '作業内容25', '工数区分名26', '定義26', '作業内容26',
        '工数区分名27', '定義27', '作業内容27', '工数区分名28', '定義28', '作業内容28',
        '工数区分名29', '定義29', '作業内容29', '工数区分名30', '定義30', '作業内容30',
        '工数区分名31', '定義31', '作業内容31', '工数区分名32', '定義32', '作業内容32',
        '工数区分名33', '定義33', '作業内容33', '工数区分名34', '定義34', '作業内容34',
        '工数区分名35', '定義35', '作業内容35', '工数区分名36', '定義36', '作業内容36',
        '工数区分名37', '定義37', '作業内容37', '工数区分名38', '定義38', '作業内容38',
        '工数区分名39', '定義39', '作業内容39', '工数区分名40', '定義40', '作業内容40',
        '工数区分名41', '定義41', '作業内容41', '工数区分名42', '定義42', '作業内容42',
        '工数区分名43', '定義43', '作業内容43', '工数区分名44', '定義44', '作業内容44',
        '工数区分名45', '定義45', '作業内容45', '工数区分名46', '定義46', '作業内容46',
        '工数区分名47', '定義47', '作業内容47', '工数区分名48', '定義48', '作業内容48',
        '工数区分名49', '定義49', '作業内容49', '工数区分名50', '定義50', '作業内容50',
        ]
    ws.append(headers)


    # Excelに書き込み(データ)
    for item in def_data:
      row = [
        item.kosu_name, item.kosu_title_1, item.kosu_division_1_1, item.kosu_division_2_1,  
        item.kosu_title_2, item.kosu_division_1_2, item.kosu_division_2_2, 
        item.kosu_title_3, item.kosu_division_1_3, item.kosu_division_2_3,  
        item.kosu_title_4, item.kosu_division_1_4, item.kosu_division_2_4,
        item.kosu_title_5, item.kosu_division_1_5, item.kosu_division_2_5,  
        item.kosu_title_6, item.kosu_division_1_6, item.kosu_division_2_6, 
        item.kosu_title_7, item.kosu_division_1_7, item.kosu_division_2_7, 
        item.kosu_title_8, item.kosu_division_1_8, item.kosu_division_2_8,
        item.kosu_title_9, item.kosu_division_1_9, item.kosu_division_2_9,  
        item.kosu_title_10, item.kosu_division_1_10, item.kosu_division_2_10,
        item.kosu_title_11, item.kosu_division_1_11, item.kosu_division_2_11,  
        item.kosu_title_12, item.kosu_division_1_12, item.kosu_division_2_12, 
        item.kosu_title_13, item.kosu_division_1_13, item.kosu_division_2_13,  
        item.kosu_title_14, item.kosu_division_1_14, item.kosu_division_2_14,
        item.kosu_title_15, item.kosu_division_1_15, item.kosu_division_2_15,  
        item.kosu_title_16, item.kosu_division_1_16, item.kosu_division_2_16, 
        item.kosu_title_17, item.kosu_division_1_17, item.kosu_division_2_17, 
        item.kosu_title_18, item.kosu_division_1_18, item.kosu_division_2_18,
        item.kosu_title_19, item.kosu_division_1_19, item.kosu_division_2_19,
        item.kosu_title_20, item.kosu_division_1_20, item.kosu_division_2_20,
        item.kosu_title_21, item.kosu_division_1_21, item.kosu_division_2_21,  
        item.kosu_title_22, item.kosu_division_1_22, item.kosu_division_2_22, 
        item.kosu_title_23, item.kosu_division_1_23, item.kosu_division_2_23,  
        item.kosu_title_24, item.kosu_division_1_24, item.kosu_division_2_24,
        item.kosu_title_25, item.kosu_division_1_25, item.kosu_division_2_25,  
        item.kosu_title_26, item.kosu_division_1_26, item.kosu_division_2_26, 
        item.kosu_title_27, item.kosu_division_1_27, item.kosu_division_2_27, 
        item.kosu_title_28, item.kosu_division_1_28, item.kosu_division_2_28,
        item.kosu_title_29, item.kosu_division_1_29, item.kosu_division_2_29,  
        item.kosu_title_30, item.kosu_division_1_30, item.kosu_division_2_30,
        item.kosu_title_31, item.kosu_division_1_31, item.kosu_division_2_31,  
        item.kosu_title_32, item.kosu_division_1_32, item.kosu_division_2_32, 
        item.kosu_title_33, item.kosu_division_1_33, item.kosu_division_2_33,  
        item.kosu_title_34, item.kosu_division_1_34, item.kosu_division_2_34,
        item.kosu_title_35, item.kosu_division_1_35, item.kosu_division_2_35,  
        item.kosu_title_36, item.kosu_division_1_36, item.kosu_division_2_36, 
        item.kosu_title_37, item.kosu_division_1_37, item.kosu_division_2_37, 
        item.kosu_title_38, item.kosu_division_1_38, item.kosu_division_2_38,
        item.kosu_title_39, item.kosu_division_1_39, item.kosu_division_2_39,
        item.kosu_title_40, item.kosu_division_1_40, item.kosu_division_2_40,
        item.kosu_title_41, item.kosu_division_1_41, item.kosu_division_2_41,  
        item.kosu_title_42, item.kosu_division_1_42, item.kosu_division_2_42, 
        item.kosu_title_43, item.kosu_division_1_43, item.kosu_division_2_43,  
        item.kosu_title_44, item.kosu_division_1_44, item.kosu_division_2_44,
        item.kosu_title_45, item.kosu_division_1_45, item.kosu_division_2_45,  
        item.kosu_title_46, item.kosu_division_1_46, item.kosu_division_2_46, 
        item.kosu_title_47, item.kosu_division_1_47, item.kosu_division_2_47, 
        item.kosu_title_48, item.kosu_division_1_48, item.kosu_division_2_48,
        item.kosu_title_49, item.kosu_division_1_49, item.kosu_division_2_49,  
        item.kosu_title_50, item.kosu_division_1_50, item.kosu_division_2_50,
        ]
      ws.append(row)


    # メモリ上にExcelファイルを作成
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # ファイル名を設定
    filename = f'工数定義区分データバックアップ_{today}.xlsx'

    # URLエンコーディングされたファイル名を生成
    quoted_filename = urllib.parse.quote(filename)

    # HttpResponseを作成してファイルをダウンロードさせる
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # Content-Dispositionヘッダーを設定
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{quoted_filename}'

    return response



  # 工数区分定義読み込み
  if 'def_load' in request.POST:
    # 工数データファイルが未選択時の処理
    if 'def_file' not in request.FILES:
      # エラーメッセージ出力
      messages.error(request, '工数区分定義ファイルが選択されていません。ERROR046')
      # このページをリダイレクト
      return redirect(to = '/administrator')


    # POSTされたファイルパスを変数に入れる
    file_path = request.FILES['def_file']

    # 一時的なファイルをサーバー上に作成
    with open('def_file_path.xlsx', 'wb+') as destination:

      # アップロードしたファイルを一時ファイルに書き込み
      for chunk in file_path.chunks():
        destination.write(chunk)

    # 指定Excelを開く
    wb = openpyxl.load_workbook('def_file_path.xlsx')
    # 書き込みシート選択
    ws = wb.worksheets[0]


    # 読み込むファイルが正しいファイルでない場合の処理
    if ws.cell(1, 1).value != '工数区分定義Ver名' or ws.cell(1, 2).value != '工数区分名1' or \
      ws.cell(1, 3).value != '定義1' or ws.cell(1, 4).value != '作業内容1' or \
      ws.cell(1, 5).value != '工数区分名2' or ws.cell(1, 6).value != '定義2' or \
      ws.cell(1, 7).value != '作業内容2' or ws.cell(1, 8).value != '工数区分名3' or \
      ws.cell(1, 9).value != '定義3' or ws.cell(1, 10).value != '作業内容3' or \
      ws.cell(1, 11).value != '工数区分名4' or ws.cell(1, 12).value != '定義4' or \
      ws.cell(1, 13).value != '作業内容4' or ws.cell(1, 14).value != '工数区分名5' or \
      ws.cell(1, 15).value != '定義5' or ws.cell(1, 16).value != '作業内容5' or \
      ws.cell(1, 17).value != '工数区分名6' or ws.cell(1, 18).value != '定義6' or \
      ws.cell(1, 19).value != '作業内容6' or ws.cell(1, 20).value != '工数区分名7' or \
      ws.cell(1, 21).value != '定義7' or ws.cell(1, 22).value != '作業内容7' or \
      ws.cell(1, 23).value != '工数区分名8' or ws.cell(1, 24).value != '定義8' or \
      ws.cell(1, 25).value != '作業内容8' or ws.cell(1, 26).value != '工数区分名9' or \
      ws.cell(1, 27).value != '定義9' or ws.cell(1, 28).value != '作業内容9' or \
      ws.cell(1, 29).value != '工数区分名10' or ws.cell(1, 30).value != '定義10' or \
      ws.cell(1, 31).value != '作業内容10' or ws.cell(1, 32).value != '工数区分名11' or \
      ws.cell(1, 33).value != '定義11' or ws.cell(1, 34).value != '作業内容11' or \
      ws.cell(1, 35).value != '工数区分名12' or ws.cell(1, 36).value != '定義12' or \
      ws.cell(1, 37).value != '作業内容12' or ws.cell(1, 38).value != '工数区分名13' or \
      ws.cell(1, 39).value != '定義13' or ws.cell(1, 40).value != '作業内容13' or \
      ws.cell(1, 41).value != '工数区分名14' or ws.cell(1, 42).value != '定義14' or \
      ws.cell(1, 43).value != '作業内容14' or ws.cell(1, 44).value != '工数区分名15' or \
      ws.cell(1, 45).value != '定義15' or ws.cell(1, 46).value != '作業内容15' or \
      ws.cell(1, 47).value != '工数区分名16' or ws.cell(1, 48).value != '定義16' or \
      ws.cell(1, 49).value != '作業内容16' or ws.cell(1, 50).value != '工数区分名17' or \
      ws.cell(1, 51).value != '定義17' or ws.cell(1, 52).value != '作業内容17' or \
      ws.cell(1, 53).value != '工数区分名18' or ws.cell(1, 54).value != '定義18' or \
      ws.cell(1, 55).value != '作業内容18' or ws.cell(1, 56).value != '工数区分名19' or \
      ws.cell(1, 57).value != '定義19' or ws.cell(1, 58).value != '作業内容19' or \
      ws.cell(1, 59).value != '工数区分名20' or ws.cell(1, 60).value != '定義20' or \
      ws.cell(1, 61).value != '作業内容20' or ws.cell(1, 62).value != '工数区分名21' or \
      ws.cell(1, 63).value != '定義21' or ws.cell(1, 64).value != '作業内容21' or \
      ws.cell(1, 65).value != '工数区分名22' or ws.cell(1, 66).value != '定義22' or \
      ws.cell(1, 67).value != '作業内容22' or ws.cell(1, 68).value != '工数区分名23' or \
      ws.cell(1, 69).value != '定義23' or ws.cell(1, 70).value != '作業内容23' or \
      ws.cell(1, 71).value != '工数区分名24' or ws.cell(1, 72).value != '定義24' or \
      ws.cell(1, 73).value != '作業内容24' or ws.cell(1, 74).value != '工数区分名25' or \
      ws.cell(1, 75).value != '定義25' or ws.cell(1, 76).value != '作業内容25' or \
      ws.cell(1, 77).value != '工数区分名26' or ws.cell(1, 78).value != '定義26' or \
      ws.cell(1, 79).value != '作業内容26' or ws.cell(1, 80).value != '工数区分名27' or \
      ws.cell(1, 81).value != '定義27' or ws.cell(1, 82).value != '作業内容27' or \
      ws.cell(1, 83).value != '工数区分名28' or ws.cell(1, 84).value != '定義28' or \
      ws.cell(1, 85).value != '作業内容28' or ws.cell(1, 86).value != '工数区分名29' or \
      ws.cell(1, 87).value != '定義29' or ws.cell(1, 88).value != '作業内容29' or \
      ws.cell(1, 89).value != '工数区分名30' or ws.cell(1, 90).value != '定義30' or \
      ws.cell(1, 91).value != '作業内容30' or ws.cell(1, 92).value != '工数区分名31' or \
      ws.cell(1, 93).value != '定義31' or ws.cell(1, 94).value != '作業内容31' or \
      ws.cell(1, 95).value != '工数区分名32' or ws.cell(1, 96).value != '定義32' or \
      ws.cell(1, 97).value != '作業内容32' or ws.cell(1, 98).value != '工数区分名33' or \
      ws.cell(1, 99).value != '定義33' or ws.cell(1, 100).value != '作業内容33' or \
      ws.cell(1, 101).value != '工数区分名34' or ws.cell(1, 102).value != '定義34' or \
      ws.cell(1, 103).value != '作業内容34' or ws.cell(1, 104).value != '工数区分名35' or \
      ws.cell(1, 105).value != '定義35' or ws.cell(1, 106).value != '作業内容35' or \
      ws.cell(1, 107).value != '工数区分名36' or ws.cell(1, 108).value != '定義36' or \
      ws.cell(1, 109).value != '作業内容36' or ws.cell(1, 110).value != '工数区分名37' or \
      ws.cell(1, 111).value != '定義37' or ws.cell(1, 112).value != '作業内容37' or \
      ws.cell(1, 113).value != '工数区分名38' or ws.cell(1, 114).value != '定義38' or \
      ws.cell(1, 115).value != '作業内容38' or ws.cell(1, 116).value != '工数区分名39' or \
      ws.cell(1, 117).value != '定義39' or ws.cell(1, 118).value != '作業内容39' or \
      ws.cell(1, 119).value != '工数区分名40' or ws.cell(1, 120).value != '定義40' or \
      ws.cell(1, 121).value != '作業内容40' or ws.cell(1, 122).value != '工数区分名41' or \
      ws.cell(1, 123).value != '定義41' or ws.cell(1, 124).value != '作業内容41' or \
      ws.cell(1, 125).value != '工数区分名42' or ws.cell(1, 126).value != '定義42' or \
      ws.cell(1, 127).value != '作業内容42' or ws.cell(1, 128).value != '工数区分名43' or \
      ws.cell(1, 129).value != '定義43' or ws.cell(1, 130).value != '作業内容43' or \
      ws.cell(1, 131).value != '工数区分名44' or ws.cell(1, 132).value != '定義44' or \
      ws.cell(1, 133).value != '作業内容44' or ws.cell(1, 134).value != '工数区分名45' or \
      ws.cell(1, 135).value != '定義45' or ws.cell(1, 136).value != '作業内容45' or \
      ws.cell(1, 137).value != '工数区分名46' or ws.cell(1, 138).value != '定義46' or \
      ws.cell(1, 139).value != '作業内容46' or ws.cell(1, 140).value != '工数区分名47' or \
      ws.cell(1, 141).value != '定義47' or ws.cell(1, 142).value != '作業内容47' or \
      ws.cell(1, 143).value != '工数区分名48' or ws.cell(1, 144).value != '定義48' or \
      ws.cell(1, 145).value != '作業内容48' or ws.cell(1, 146).value != '工数区分名49' or \
      ws.cell(1, 147).value != '定義49' or ws.cell(1, 148).value != '作業内容49' or \
      ws.cell(1, 149).value != '工数区分名50' or ws.cell(1, 150).value != '定義50' or \
      ws.cell(1, 151).value != '作業内容50':
      # エラーメッセージ出力
      messages.error(request, 'ロードしようとしたファイルは工数区分定義情報バックアップではありません。ERROR051')
      # このページをリダイレクト
      return redirect(to = '/administrator')


    # レコード数取得
    data_num = ws.max_row


    # Excelからデータを読み込むループ
    for i in range(1, data_num):
      # 読み込み予定データと同一の定義の名前のデータが存在するか確認
      def_data_filter = kosu_division.objects.filter(kosu_name = ws.cell(row = i + 1, column = 1).value)
      # 読み込み予定データと同一の定義の名前のデータが存在する場合の処理
      if def_data_filter.count() != 0:
        # 読み込み予定データと同一の定義の名前のデータを取得
        def_data_get = kosu_division.objects.get(kosu_name = ws.cell(row = i + 1, column = 1).value)
        # 読み込み予定データと同一の定義の名前のデータを削除
        def_data_get.delete()

      # Excelからデータを読み込む
      new_data = kosu_division(kosu_name = ws.cell(row = i + 1, column = 1).value, \
                               kosu_title_1 = ws.cell(row = i + 1, column = 2).value, \
                               kosu_division_1_1 = ws.cell(row = i + 1, column = 3).value, \
                               kosu_division_2_1 = ws.cell(row = i + 1, column = 4).value, \
                               kosu_title_2 = ws.cell(row = i + 1, column = 5).value, \
                               kosu_division_1_2 = ws.cell(row = i + 1, column = 6).value, \
                               kosu_division_2_2 = ws.cell(row = i + 1, column = 7).value, \
                               kosu_title_3 = ws.cell(row = i + 1, column = 8).value, \
                               kosu_division_1_3 = ws.cell(row = i + 1, column = 9).value, \
                               kosu_division_2_3 = ws.cell(row = i + 1, column = 10).value, \
                               kosu_title_4 = ws.cell(row = i + 1, column = 11).value, \
                               kosu_division_1_4 = ws.cell(row = i + 1, column = 12).value, \
                               kosu_division_2_4 = ws.cell(row = i + 1, column = 13).value, \
                               kosu_title_5 = ws.cell(row = i + 1, column = 14).value, \
                               kosu_division_1_5 = ws.cell(row = i + 1, column = 15).value, \
                               kosu_division_2_5 = ws.cell(row = i + 1, column = 16).value, \
                               kosu_title_6 = ws.cell(row = i + 1, column = 17).value, \
                               kosu_division_1_6 = ws.cell(row = i + 1, column = 18).value, \
                               kosu_division_2_6 = ws.cell(row = i + 1, column = 19).value, \
                               kosu_title_7 = ws.cell(row = i + 1, column = 20).value, \
                               kosu_division_1_7 = ws.cell(row = i + 1, column = 21).value, \
                               kosu_division_2_7 = ws.cell(row = i + 1, column = 22).value, \
                               kosu_title_8 = ws.cell(row = i + 1, column = 23).value, \
                               kosu_division_1_8 = ws.cell(row = i + 1, column = 24).value, \
                               kosu_division_2_8 = ws.cell(row = i + 1, column = 25).value, \
                               kosu_title_9 = ws.cell(row = i + 1, column = 26).value, \
                               kosu_division_1_9 = ws.cell(row = i + 1, column = 27).value, \
                               kosu_division_2_9 = ws.cell(row = i + 1, column = 28).value, \
                               kosu_title_10 = ws.cell(row = i + 1, column = 29).value, \
                               kosu_division_1_10 = ws.cell(row = i + 1, column = 30).value, \
                               kosu_division_2_10 = ws.cell(row = i + 1, column = 31).value, \
                               kosu_title_11 = ws.cell(row = i + 1, column = 32).value, \
                               kosu_division_1_11 = ws.cell(row = i + 1, column = 33).value, \
                               kosu_division_2_11 = ws.cell(row = i + 1, column = 34).value, \
                               kosu_title_12 = ws.cell(row = i + 1, column = 35).value, \
                               kosu_division_1_12 = ws.cell(row = i + 1, column = 36).value, \
                               kosu_division_2_12 = ws.cell(row = i + 1, column = 37).value, \
                               kosu_title_13 = ws.cell(row = i + 1, column = 38).value, \
                               kosu_division_1_13 = ws.cell(row = i + 1, column = 39).value, \
                               kosu_division_2_13 = ws.cell(row = i + 1, column = 40).value, \
                               kosu_title_14 = ws.cell(row = i + 1, column = 41).value, \
                               kosu_division_1_14 = ws.cell(row = i + 1, column = 42).value, \
                               kosu_division_2_14 = ws.cell(row = i + 1, column = 43).value, \
                               kosu_title_15 = ws.cell(row = i + 1, column = 44).value, \
                               kosu_division_1_15 = ws.cell(row = i + 1, column = 45).value, \
                               kosu_division_2_15 = ws.cell(row = i + 1, column = 46).value, \
                               kosu_title_16 = ws.cell(row = i + 1, column = 47).value, \
                               kosu_division_1_16 = ws.cell(row = i + 1, column = 48).value, \
                               kosu_division_2_16 = ws.cell(row = i + 1, column = 49).value, \
                               kosu_title_17 = ws.cell(row = i + 1, column = 50).value, \
                               kosu_division_1_17 = ws.cell(row = i + 1, column = 51).value, \
                               kosu_division_2_17 = ws.cell(row = i + 1, column = 52).value, \
                               kosu_title_18 = ws.cell(row = i + 1, column = 53).value, \
                               kosu_division_1_18 = ws.cell(row = i + 1, column = 54).value, \
                               kosu_division_2_18 = ws.cell(row = i + 1, column = 55).value, \
                               kosu_title_19 = ws.cell(row = i + 1, column = 56).value, \
                               kosu_division_1_19 = ws.cell(row = i + 1, column = 57).value, \
                               kosu_division_2_19 = ws.cell(row = i + 1, column = 58).value, \
                               kosu_title_20 = ws.cell(row = i + 1, column = 59).value, \
                               kosu_division_1_20 = ws.cell(row = i + 1, column = 60).value, \
                               kosu_division_2_20 = ws.cell(row = i + 1, column = 61).value, \
                               kosu_title_21 = ws.cell(row = i + 1, column = 62).value, \
                               kosu_division_1_21 = ws.cell(row = i + 1, column = 63).value, \
                               kosu_division_2_21 = ws.cell(row = i + 1, column = 64).value, \
                               kosu_title_22 = ws.cell(row = i + 1, column = 65).value, \
                               kosu_division_1_22 = ws.cell(row = i + 1, column = 66).value, \
                               kosu_division_2_22 = ws.cell(row = i + 1, column = 67).value, \
                               kosu_title_23 = ws.cell(row = i + 1, column = 68).value, \
                               kosu_division_1_23 = ws.cell(row = i + 1, column = 69).value, \
                               kosu_division_2_23 = ws.cell(row = i + 1, column = 70).value, \
                               kosu_title_24 = ws.cell(row = i + 1, column = 71).value, \
                               kosu_division_1_24 = ws.cell(row = i + 1, column = 72).value, \
                               kosu_division_2_24 = ws.cell(row = i + 1, column = 73).value, \
                               kosu_title_25 = ws.cell(row = i + 1, column = 74).value, \
                               kosu_division_1_25 = ws.cell(row = i + 1, column = 75).value, \
                               kosu_division_2_25 = ws.cell(row = i + 1, column = 76).value, \
                               kosu_title_26 = ws.cell(row = i + 1, column = 77).value, \
                               kosu_division_1_26 = ws.cell(row = i + 1, column = 78).value, \
                               kosu_division_2_26 = ws.cell(row = i + 1, column = 79).value, \
                               kosu_title_27 = ws.cell(row = i + 1, column = 80).value, \
                               kosu_division_1_27 = ws.cell(row = i + 1, column = 81).value, \
                               kosu_division_2_27 = ws.cell(row = i + 1, column = 82).value, \
                               kosu_title_28 = ws.cell(row = i + 1, column = 83).value, \
                               kosu_division_1_28 = ws.cell(row = i + 1, column = 84).value, \
                               kosu_division_2_28 = ws.cell(row = i + 1, column = 85).value, \
                               kosu_title_29 = ws.cell(row = i + 1, column = 86).value, \
                               kosu_division_1_29 = ws.cell(row = i + 1, column = 87).value, \
                               kosu_division_2_29 = ws.cell(row = i + 1, column = 88).value, \
                               kosu_title_30 = ws.cell(row = i + 1, column = 89).value, \
                               kosu_division_1_30 = ws.cell(row = i + 1, column = 90).value, \
                               kosu_division_2_30 = ws.cell(row = i + 1, column = 91).value, \
                               kosu_title_31 = ws.cell(row = i + 1, column = 92).value, \
                               kosu_division_1_31 = ws.cell(row = i + 1, column = 93).value, \
                               kosu_division_2_31 = ws.cell(row = i + 1, column = 94).value, \
                               kosu_title_32 = ws.cell(row = i + 1, column = 95).value, \
                               kosu_division_1_32 = ws.cell(row = i + 1, column = 96).value, \
                               kosu_division_2_32 = ws.cell(row = i + 1, column = 97).value, \
                               kosu_title_33 = ws.cell(row = i + 1, column = 98).value, \
                               kosu_division_1_33 = ws.cell(row = i + 1, column = 99).value, \
                               kosu_division_2_33 = ws.cell(row = i + 1, column = 100).value, \
                               kosu_title_34 = ws.cell(row = i + 1, column = 101).value, \
                               kosu_division_1_34 = ws.cell(row = i + 1, column = 102).value, \
                               kosu_division_2_34 = ws.cell(row = i + 1, column = 103).value, \
                               kosu_title_35 = ws.cell(row = i + 1, column = 104).value, \
                               kosu_division_1_35 = ws.cell(row = i + 1, column = 105).value, \
                               kosu_division_2_35 = ws.cell(row = i + 1, column = 106).value, \
                               kosu_title_36 = ws.cell(row = i + 1, column = 107).value, \
                               kosu_division_1_36 = ws.cell(row = i + 1, column = 108).value, \
                               kosu_division_2_36 = ws.cell(row = i + 1, column = 109).value, \
                               kosu_title_37 = ws.cell(row = i + 1, column = 110).value, \
                               kosu_division_1_37 = ws.cell(row = i + 1, column = 111).value, \
                               kosu_division_2_37 = ws.cell(row = i + 1, column = 112).value, \
                               kosu_title_38 = ws.cell(row = i + 1, column = 113).value, \
                               kosu_division_1_38 = ws.cell(row = i + 1, column = 114).value, \
                               kosu_division_2_38 = ws.cell(row = i + 1, column = 115).value, \
                               kosu_title_39 = ws.cell(row = i + 1, column = 116).value, \
                               kosu_division_1_39 = ws.cell(row = i + 1, column = 117).value, \
                               kosu_division_2_39 = ws.cell(row = i + 1, column = 118).value, \
                               kosu_title_40 = ws.cell(row = i + 1, column = 119).value, \
                               kosu_division_1_40 = ws.cell(row = i + 1, column = 120).value, \
                               kosu_division_2_40 = ws.cell(row = i + 1, column = 121).value, \
                               kosu_title_41 = ws.cell(row = i + 1, column = 122).value, \
                               kosu_division_1_41 = ws.cell(row = i + 1, column = 123).value, \
                               kosu_division_2_41 = ws.cell(row = i + 1, column = 124).value, \
                               kosu_title_42 = ws.cell(row = i + 1, column = 125).value, \
                               kosu_division_1_42 = ws.cell(row = i + 1, column = 126).value, \
                               kosu_division_2_42 = ws.cell(row = i + 1, column = 127).value, \
                               kosu_title_43 = ws.cell(row = i + 1, column = 128).value, \
                               kosu_division_1_43 = ws.cell(row = i + 1, column = 129).value, \
                               kosu_division_2_43 = ws.cell(row = i + 1, column = 130).value, \
                               kosu_title_44 = ws.cell(row = i + 1, column = 131).value, \
                               kosu_division_1_44 = ws.cell(row = i + 1, column = 132).value, \
                               kosu_division_2_44 = ws.cell(row = i + 1, column = 133).value, \
                               kosu_title_45 = ws.cell(row = i + 1, column = 134).value, \
                               kosu_division_1_45 = ws.cell(row = i + 1, column = 135).value, \
                               kosu_division_2_45 = ws.cell(row = i + 1, column = 136).value, \
                               kosu_title_46 = ws.cell(row = i + 1, column = 137).value, \
                               kosu_division_1_46 = ws.cell(row = i + 1, column = 138).value, \
                               kosu_division_2_46 = ws.cell(row = i + 1, column = 139).value, \
                               kosu_title_47 = ws.cell(row = i + 1, column = 140).value, \
                               kosu_division_1_47 = ws.cell(row = i + 1, column = 141).value, \
                               kosu_division_2_47 = ws.cell(row = i + 1, column = 142).value, \
                               kosu_title_48 = ws.cell(row = i + 1, column = 143).value, \
                               kosu_division_1_48 = ws.cell(row = i + 1, column = 144).value, \
                               kosu_division_2_48 = ws.cell(row = i + 1, column = 145).value, \
                               kosu_title_49 = ws.cell(row = i + 1, column = 146).value, \
                               kosu_division_1_49 = ws.cell(row = i + 1, column = 147).value, \
                               kosu_division_2_49 = ws.cell(row = i + 1, column = 148).value, \
                               kosu_title_50 = ws.cell(row = i + 1, column = 149).value, \
                               kosu_division_1_50 = ws.cell(row = i + 1, column = 150).value, \
                               kosu_division_2_50 = ws.cell(row = i + 1, column = 151).value)

      new_data.save()

    # 一時ファイル削除
    os.remove('def_file_path.xlsx')

  

  # お問い合わせバックアップ処理
  if 'inquiry_backup' in request.POST:
    # 今日の日付取得
    today = datetime.date.today().strftime('%Y%m%d')
    # 新しいExcelブック作成
    wb = openpyxl.Workbook()
    # 書き込みシート選択
    ws = wb.active
    # お問い合わせデータ取得
    data = inquiry_data.objects.all()

    # Excelに書き込み(項目名)
    headers = [
        '従業員番号', 
        '氏名', 
        '内容選択', 
        '問い合わせ', 
        '回答'
        ]
    ws.append(headers)


    # Excelに書き込み(データ)
    for item in data:
      row = [
        item.employee_no2, 
        str(item.name), 
        item.content_choice, 
        item.inquiry, item.answer
        ]
      ws.append(row)

    # メモリ上にExcelファイルを作成
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # ファイル名を設定
    filename = f'お問い合わせデータバックアップ_{today}.xlsx'

    # URLエンコーディングされたファイル名を生成
    quoted_filename = urllib.parse.quote(filename)

    # HttpResponseを作成してファイルをダウンロードさせる
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # Content-Dispositionヘッダーを設定
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{quoted_filename}'
    
    return response
  


  # お問い合わせ読み込み
  if 'inquiry_load' in request.POST:
    # 工数データファイルが未選択時の処理
    if 'inquiry_file' not in request.FILES:
      # エラーメッセージ出力
      messages.error(request, 'お問い合わせファイルが選択されていません。ERROR043')
      # このページをリダイレクト
      return redirect(to = '/administrator')


    # POSTされたファイルパスを変数に入れる
    file_path = request.FILES['inquiry_file']
    # 一時的なファイルをサーバー上に作成
    with open('inquiry_file_path.xlsx', 'wb+') as destination:
      # アップロードしたファイルを一時ファイルに書き込み
      for chunk in file_path.chunks():
        destination.write(chunk)
  
    # 指定Excelを開く
    wb = openpyxl.load_workbook('inquiry_file_path.xlsx')
    # 書き込みシート選択
    ws = wb.worksheets[0]


    # 読み込むファイルが正しいファイルでない場合の処理
    if ws.cell(1, 1).value != '従業員番号' or ws.cell(1, 2).value != '氏名' or \
      ws.cell(1, 3).value != '内容選択' or ws.cell(1, 4).value != '問い合わせ' or \
      ws.cell(1, 5).value != '回答':
      # エラーメッセージ出力
      messages.error(request, 'ロードしようとしたファイルはお問い合わせ情報バックアップではありません。ERROR030')
      # このページをリダイレクト
      return redirect(to = '/administrator')


    # レコード数取得
    data_num = ws.max_row
    # 全てのお問い合わせデータを取得
    inquiry_data_get = inquiry_data.objects.all()
    # 読お問い合わせデータを削除
    inquiry_data_get.delete()


    # Excelからデータを読み込むループ
    for i in range(1, data_num):
      # 人員データインスタンス取得
      member_instance = member.objects.get(name = ws.cell(row = i + 1, column = 2).value)
      # Excelからデータを読み込む
      new_data = inquiry_data(employee_no2 = ws.cell(row = i + 1, column = 1).value, \
                              name = member_instance, \
                              content_choice = ws.cell(row = i + 1, column = 3).value, \
                              inquiry = ws.cell(row = i + 1, column = 4).value, \
                              answer = ws.cell(row = i + 1, column = 5).value)

      new_data.save()

    # 一時ファイル削除
    os.remove('inquiry_file_path.xlsx')



  # 設定情報バックアップ処理
  if 'setting_backup' in request.POST:
    # 今日の日付取得
    today = today = datetime.date.today().strftime('%Y%m%d')
    # 新しいExcelブック作成
    wb = openpyxl.Workbook()
    # 書き込みシート選択
    ws = wb.active

    # 設定データ取得
    setting_data = administrator_data.objects.all()

    # Excelに書き込み(項目名)
    headers = [
        '一覧表示項目数', 
        '問い合わせ担当者従業員番号1',
        '問い合わせ担当者従業員番号2',
        '問い合わせ担当者従業員番号3',
        ]
    ws.append(headers)


    # Excelに書き込み(データ)
    for item in setting_data:
      row = [
        item.menu_row,
        item.administrator_employee_no1,
        item.administrator_employee_no2,
        item.administrator_employee_no3,
        ]
      ws.append(row)


    # メモリ上にExcelファイルを作成
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # ファイル名を設定
    filename = f'管理者設定_{today}.xlsx'

    # URLエンコーディングされたファイル名を生成
    quoted_filename = urllib.parse.quote(filename)
    
    # HttpResponseを作成してファイルをダウンロードさせる
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # Content-Dispositionヘッダーを設定
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{quoted_filename}'
    
    return response
  


  # 設定情報読み込み
  if 'setting_load' in request.POST:
    # 工数データファイルが未選択時の処理
    if 'setting_file' not in request.FILES:
      # エラーメッセージ出力
      messages.error(request, '管理者設定ファイルが選択されていません。ERROR047')
      # このページをリダイレクト
      return redirect(to = '/administrator')


    # POSTされたファイルパスを変数に入れる
    file_path = request.FILES['setting_file']


    # 一時的なファイルをサーバー上に作成
    with open('setting_file_path.xlsx', 'wb+') as destination:

      # アップロードしたファイルを一時ファイルに書き込み
      for chunk in file_path.chunks():
        destination.write(chunk)

    # 指定Excelを開く
    wb = openpyxl.load_workbook('setting_file_path.xlsx')
    # 書き込みシート選択
    ws = wb.worksheets[0]


    # 読み込むファイルが正しいファイルでない場合の処理
    if ws.cell(1, 1).value != '一覧表示項目数' or ws.cell(1, 2).value != '問い合わせ担当者従業員番号1' or \
      ws.cell(1, 3).value != '問い合わせ担当者従業員番号2' or ws.cell(1, 4).value != '問い合わせ担当者従業員番号4':
      # エラーメッセージ出力
      messages.error(request, 'ロードしようとしたファイルは設定情報バックアップではありません。ERROR052')
      # このページをリダイレクト
      return redirect(to = '/administrator')


    # レコード数取得
    data_num = ws.max_row


    # 管理者設定データにレコードがある場合の処理
    if administrator_data.objects.exists():
      # 管理者設定データ取得
      setting_obj_get = administrator_data.objects.all()
      # 取得した管理者設定データを消す
      setting_obj_get.delete()


    # Excelからデータを読み込むループ
    for i in range(1, data_num):
      # Excelからデータを読み込み
      new_data = administrator_data(menu_row = ws.cell(row = i + 1, column = 1).value, \
                                    administrator_employee_no1 = ws.cell(row = i + 1, column = 2).value, \
                                    administrator_employee_no2 = ws.cell(row = i + 1, column = 3).value, \
                                    administrator_employee_no3 = ws.cell(row = i + 1, column = 4).value)
      new_data.save()

    # 一時ファイル削除
    os.remove('member_file_path.xlsx')


    # このページを読み直す
    return redirect(to = '/administrator')



  # HTMLに渡す辞書
  context = {
    'title' : '管理者MENU',
    'form' : form,
    'load_form' : load_form,
    }
  


  # 指定したHTMLに辞書を渡して表示を完成させる
  return render(request, 'kosu/administrator_menu.html', context)





#--------------------------------------------------------------------------------------------------------





# ヘルプ画面定義
def help(request):

  # ルートディレクトリを取得
  BASE_DIR = Path(__file__).resolve().parent.parent
  # 環境変数ファイルを読み込む
  env = environ.Env()
  env.read_env(os.path.join(BASE_DIR, '.env'))


  # フォームを定義
  form = uploadForm(request.POST)



  # GET時の処理
  if (request.method == 'GET'):
    # ファイルロード欄非表示
    display = False



  # 復帰パスワード入力時処理
  if 'help_button' in request.POST:
    # パスワード判定
    if request.POST['help_path'] ==  env('HELP_PATH'):
      display = True
    else:
      display = False
      messages.error(request, 'パスワードが違います。ERROR042')
   


  # データ読み込み
  if 'data_load' in request.POST:
    
    # データ読み込み欄表示
    display = True


    # 人員ファイルが未選択時の処理
    if 'member_file' not in request.FILES:

      # エラーメッセージ出力
      messages.error(request, '人員ファイルが選択されていません。ERROR024')

      # このページをリダイレクト
      return redirect(to = '/help')


    # 工数区分定義ファイルが未選択時の処理
    if 'def_file' not in request.FILES:

      # エラーメッセージ出力
      messages.error(request, '工数区分定義ファイルが選択されていません。ERROR025')

      # このページをリダイレクト
      return redirect(to = '/help')


    # 工数区分定義ファイルが未選択時の処理
    if 'setting_file' not in request.FILES:

      # エラーメッセージ出力
      messages.error(request, '管理者設定ファイルが選択されていません。ERROR039')

      # このページをリダイレクト
      return redirect(to = '/help')


    # 人員ファイル定義
    uploaded_file = request.FILES['member_file']

    # 一時的なファイルをサーバー上に作成
    with open('member_file_path.xlsx', 'wb+') as destination:

      # アップロードしたファイルを一時ファイルに書き込み
      for chunk in uploaded_file.chunks():
        destination.write(chunk)

    # 一時ファイルを開く
    wb = openpyxl.load_workbook('member_file_path.xlsx')

    # 一番最初のシートを指定
    ws = wb.worksheets[0]

    # レコード数取得
    data_num = ws.max_row


    # 人員データにレコードがある場合の処理
    if member.objects.exists():

      # Excelからデータを読み込むループ
      for i in range(1, data_num):

        # 人員データに指定従業員番号があるか確認
        member_obj_filter = member.objects.filter(employee_no = ws.cell(row = i + 1, column = 1).value)

        # 人員データに指定従業員番号がある場合の処理
        if member_obj_filter.count() != 0:

          # 指定従業員番号の人員データ取得
          member_obj_get = member.objects.get(employee_no = ws.cell(row = i + 1, column = 1).value)
          # 取得した人員データを消す
          member_obj_get.delete()

        # Excelからデータを読み込み
        new_data = member(employee_no = ws.cell(row = i + 1, column = 1).value, \
                          name = ws.cell(row = i + 1, column = 2).value, \
                          shop = ws.cell(row = i + 1, column = 3).value, \
                          authority = ws.cell(row = i + 1, column = 4).value, \
                          administrator = ws.cell(row = i + 1, column = 5).value, \
                          break_time1 = ws.cell(row = i + 1, column = 6).value, \
                          break_time1_over1 = ws.cell(row = i + 1, column = 7).value, \
                          break_time1_over2 = ws.cell(row = i + 1, column = 8).value, \
                          break_time1_over3 = ws.cell(row = i + 1, column = 9).value, \
                          break_time2 = ws.cell(row = i + 1, column = 10).value, \
                          break_time2_over1 = ws.cell(row = i + 1, column = 11).value, \
                          break_time2_over2 = ws.cell(row = i + 1, column = 12).value, \
                          break_time2_over3 = ws.cell(row = i + 1, column = 13).value, \
                          break_time3 = ws.cell(row = i + 1, column = 14).value, \
                          break_time3_over1 = ws.cell(row = i + 1, column = 15).value, \
                          break_time3_over2 = ws.cell(row = i + 1, column = 16).value, \
                          break_time3_over3 = ws.cell(row = i + 1, column = 17).value, \
                          break_time4 = ws.cell(row = i + 1, column = 18).value, \
                          break_time4_over1 = ws.cell(row = i + 1, column = 19).value, \
                          break_time4_over2 = ws.cell(row = i + 1, column = 20).value, \
                          break_time4_over3 = ws.cell(row = i + 1, column = 21).value)                           

        new_data.save()


    # 人員データにレコードがない場合の処理
    else:

      # Excelからデータを読み込むループ
      for i in range(1, data_num):

        # Excelからデータを読み込み
        new_data = member(employee_no = ws.cell(row = i + 1, column = 1).value, \
                          name = ws.cell(row = i + 1, column = 2).value, \
                          shop = ws.cell(row = i + 1, column = 3).value, \
                          authority = ws.cell(row = i + 1, column = 4).value, \
                          administrator = ws.cell(row = i + 1, column = 5).value, \
                          break_time1 = ws.cell(row = i + 1, column = 6).value, \
                          break_time1_over1 = ws.cell(row = i + 1, column = 7).value, \
                          break_time1_over2 = ws.cell(row = i + 1, column = 8).value, \
                          break_time1_over3 = ws.cell(row = i + 1, column = 9).value, \
                          break_time2 = ws.cell(row = i + 1, column = 10).value, \
                          break_time2_over1 = ws.cell(row = i + 1, column = 11).value, \
                          break_time2_over2 = ws.cell(row = i + 1, column = 12).value, \
                          break_time2_over3 = ws.cell(row = i + 1, column = 13).value, \
                          break_time3 = ws.cell(row = i + 1, column = 14).value, \
                          break_time3_over1 = ws.cell(row = i + 1, column = 15).value, \
                          break_time3_over2 = ws.cell(row = i + 1, column = 16).value, \
                          break_time3_over3 = ws.cell(row = i + 1, column = 17).value, \
                          break_time4 = ws.cell(row = i + 1, column = 18).value, \
                          break_time4_over1 = ws.cell(row = i + 1, column = 19).value, \
                          break_time4_over2 = ws.cell(row = i + 1, column = 20).value, \
                          break_time4_over3 = ws.cell(row = i + 1, column = 21).value)                           

        new_data.save()


    # 一時ファイル削除
    os.remove('member_file_path.xlsx')



    # 工数定義区分ファイル定義
    uploaded_file = request.FILES['def_file']

    # 一時的なファイルをサーバー上に作成
    with open('def_file_path.xlsx', 'wb+') as destination:

      # アップロードしたファイルを一時ファイルに書き込み
      for chunk in uploaded_file.chunks():
        destination.write(chunk)

    # 一時ファイルを開く
    wb1 = openpyxl.load_workbook('def_file_path.xlsx')

    # 一番最初のシートを指定
    ws1 = wb1.worksheets[0]

    # レコード数取得
    data_num = ws1.max_row


    # 工数定義区分データにレコードがある場合の処理
    if kosu_division.objects.exists():

      # Excelからデータを読み込み
      for i in range(1, data_num):

        # 工数定義区分データに指定従業員番号があるか確認
        def_obj_filter = kosu_division.objects.filter(kosu_name = ws1.cell(row = i + 1, column = 1).value)

        # 工数定義区分データに指定従業員番号がある場合の処理
        if def_obj_filter.count() != 0:

          # 指定従工数定義区分データ取得
          def_obj_get = kosu_division.objects.get(kosu_name = ws1.cell(row = i + 1, column = 1).value)
          # 取得した工数定義区分データを消す
          def_obj_get.delete()

        new_data1 = kosu_division(kosu_name = ws1.cell(row = i + 1, column = 1).value, \
                                  kosu_title_1 = ws1.cell(row = i + 1, column = 2).value, \
                                  kosu_division_1_1 = ws1.cell(row = i + 1, column = 3).value, \
                                  kosu_division_2_1 = ws1.cell(row = i + 1, column = 4).value, \
                                  kosu_title_2 = ws1.cell(row = i + 1, column = 5).value, \
                                  kosu_division_1_2 = ws1.cell(row = i + 1, column = 6).value, \
                                  kosu_division_2_2 = ws1.cell(row = i + 1, column = 7).value, \
                                  kosu_title_3 = ws1.cell(row = i + 1, column = 8).value, \
                                  kosu_division_1_3 = ws1.cell(row = i + 1, column = 9).value, \
                                  kosu_division_2_3 = ws1.cell(row = i + 1, column = 10).value, \
                                  kosu_title_4 = ws1.cell(row = i + 1, column = 11).value, \
                                  kosu_division_1_4 = ws1.cell(row = i + 1, column = 12).value, \
                                  kosu_division_2_4 = ws1.cell(row = i + 1, column = 13).value, \
                                  kosu_title_5 = ws1.cell(row = i + 1, column = 14).value, \
                                  kosu_division_1_5 = ws1.cell(row = i + 1, column = 15).value, \
                                  kosu_division_2_5 = ws1.cell(row = i + 1, column = 16).value, \
                                  kosu_title_6 = ws1.cell(row = i + 1, column = 17).value, \
                                  kosu_division_1_6 = ws1.cell(row = i + 1, column = 18).value, \
                                  kosu_division_2_6 = ws1.cell(row = i + 1, column = 19).value, \
                                  kosu_title_7 = ws1.cell(row = i + 1, column = 20).value, \
                                  kosu_division_1_7 = ws1.cell(row = i + 1, column = 21).value, \
                                  kosu_division_2_7 = ws1.cell(row = i + 1, column = 22).value, \
                                  kosu_title_8 = ws1.cell(row = i + 1, column = 23).value, \
                                  kosu_division_1_8 = ws1.cell(row = i + 1, column = 24).value, \
                                  kosu_division_2_8 = ws1.cell(row = i + 1, column = 25).value, \
                                  kosu_title_9 = ws1.cell(row = i + 1, column = 26).value, \
                                  kosu_division_1_9 = ws1.cell(row = i + 1, column = 27).value, \
                                  kosu_division_2_9 = ws1.cell(row = i + 1, column = 28).value, \
                                  kosu_title_10 = ws1.cell(row = i + 1, column = 29).value, \
                                  kosu_division_1_10 = ws1.cell(row = i + 1, column = 30).value, \
                                  kosu_division_2_10 = ws1.cell(row = i + 1, column = 31).value, \
                                  kosu_title_11 = ws1.cell(row = i + 1, column = 32).value, \
                                  kosu_division_1_11 = ws1.cell(row = i + 1, column = 33).value, \
                                  kosu_division_2_11 = ws1.cell(row = i + 1, column = 34).value, \
                                  kosu_title_12 = ws1.cell(row = i + 1, column = 35).value, \
                                  kosu_division_1_12 = ws1.cell(row = i + 1, column = 36).value, \
                                  kosu_division_2_12 = ws1.cell(row = i + 1, column = 37).value, \
                                  kosu_title_13 = ws1.cell(row = i + 1, column = 38).value, \
                                  kosu_division_1_13 = ws1.cell(row = i + 1, column = 39).value, \
                                  kosu_division_2_13 = ws1.cell(row = i + 1, column = 40).value, \
                                  kosu_title_14 = ws1.cell(row = i + 1, column = 41).value, \
                                  kosu_division_1_14 = ws1.cell(row = i + 1, column = 42).value, \
                                  kosu_division_2_14 = ws1.cell(row = i + 1, column = 43).value, \
                                  kosu_title_15 = ws1.cell(row = i + 1, column = 44).value, \
                                  kosu_division_1_15 = ws1.cell(row = i + 1, column = 45).value, \
                                  kosu_division_2_15 = ws1.cell(row = i + 1, column = 46).value, \
                                  kosu_title_16 = ws1.cell(row = i + 1, column = 47).value, \
                                  kosu_division_1_16 = ws1.cell(row = i + 1, column = 48).value, \
                                  kosu_division_2_16 = ws1.cell(row = i + 1, column = 49).value, \
                                  kosu_title_17 = ws1.cell(row = i + 1, column = 50).value, \
                                  kosu_division_1_17 = ws1.cell(row = i + 1, column = 51).value, \
                                  kosu_division_2_17 = ws1.cell(row = i + 1, column = 52).value, \
                                  kosu_title_18 = ws1.cell(row = i + 1, column = 53).value, \
                                  kosu_division_1_18 = ws1.cell(row = i + 1, column = 54).value, \
                                  kosu_division_2_18 = ws1.cell(row = i + 1, column = 55).value, \
                                  kosu_title_19 = ws1.cell(row = i + 1, column = 56).value, \
                                  kosu_division_1_19 = ws1.cell(row = i + 1, column = 57).value, \
                                  kosu_division_2_19 = ws1.cell(row = i + 1, column = 58).value, \
                                  kosu_title_20 = ws1.cell(row = i + 1, column = 59).value, \
                                  kosu_division_1_20 = ws1.cell(row = i + 1, column = 60).value, \
                                  kosu_division_2_20 = ws1.cell(row = i + 1, column = 61).value, \
                                  kosu_title_21 = ws1.cell(row = i + 1, column = 62).value, \
                                  kosu_division_1_21 = ws1.cell(row = i + 1, column = 63).value, \
                                  kosu_division_2_21 = ws1.cell(row = i + 1, column = 64).value, \
                                  kosu_title_22 = ws1.cell(row = i + 1, column = 65).value, \
                                  kosu_division_1_22 = ws1.cell(row = i + 1, column = 66).value, \
                                  kosu_division_2_22 = ws1.cell(row = i + 1, column = 67).value, \
                                  kosu_title_23 = ws1.cell(row = i + 1, column = 68).value, \
                                  kosu_division_1_23 = ws1.cell(row = i + 1, column = 69).value, \
                                  kosu_division_2_23 = ws1.cell(row = i + 1, column = 70).value, \
                                  kosu_title_24 = ws1.cell(row = i + 1, column = 71).value, \
                                  kosu_division_1_24 = ws1.cell(row = i + 1, column = 72).value, \
                                  kosu_division_2_24 = ws1.cell(row = i + 1, column = 73).value, \
                                  kosu_title_25 = ws1.cell(row = i + 1, column = 74).value, \
                                  kosu_division_1_25 = ws1.cell(row = i + 1, column = 75).value, \
                                  kosu_division_2_25 = ws1.cell(row = i + 1, column = 76).value, \
                                  kosu_title_26 = ws1.cell(row = i + 1, column = 77).value, \
                                  kosu_division_1_26 = ws1.cell(row = i + 1, column = 78).value, \
                                  kosu_division_2_26 = ws1.cell(row = i + 1, column = 79).value, \
                                  kosu_title_27 = ws1.cell(row = i + 1, column = 80).value, \
                                  kosu_division_1_27 = ws1.cell(row = i + 1, column = 81).value, \
                                  kosu_division_2_27 = ws1.cell(row = i + 1, column = 82).value, \
                                  kosu_title_28 = ws1.cell(row = i + 1, column = 83).value, \
                                  kosu_division_1_28 = ws1.cell(row = i + 1, column = 84).value, \
                                  kosu_division_2_28 = ws1.cell(row = i + 1, column = 85).value, \
                                  kosu_title_29 = ws1.cell(row = i + 1, column = 86).value, \
                                  kosu_division_1_29 = ws1.cell(row = i + 1, column = 87).value, \
                                  kosu_division_2_29 = ws1.cell(row = i + 1, column = 88).value, \
                                  kosu_title_30 = ws1.cell(row = i + 1, column = 89).value, \
                                  kosu_division_1_30 = ws1.cell(row = i + 1, column = 90).value, \
                                  kosu_division_2_30 = ws1.cell(row = i + 1, column = 91).value, \
                                  kosu_title_31 = ws1.cell(row = i + 1, column = 92).value, \
                                  kosu_division_1_31 = ws1.cell(row = i + 1, column = 93).value, \
                                  kosu_division_2_31 = ws1.cell(row = i + 1, column = 94).value, \
                                  kosu_title_32 = ws1.cell(row = i + 1, column = 95).value, \
                                  kosu_division_1_32 = ws1.cell(row = i + 1, column = 96).value, \
                                  kosu_division_2_32 = ws1.cell(row = i + 1, column = 97).value, \
                                  kosu_title_33 = ws1.cell(row = i + 1, column = 98).value, \
                                  kosu_division_1_33 = ws1.cell(row = i + 1, column = 99).value, \
                                  kosu_division_2_33 = ws1.cell(row = i + 1, column = 100).value, \
                                  kosu_title_34 = ws1.cell(row = i + 1, column = 101).value, \
                                  kosu_division_1_34 = ws1.cell(row = i + 1, column = 102).value, \
                                  kosu_division_2_34 = ws1.cell(row = i + 1, column = 103).value, \
                                  kosu_title_35 = ws1.cell(row = i + 1, column = 104).value, \
                                  kosu_division_1_35 = ws1.cell(row = i + 1, column = 105).value, \
                                  kosu_division_2_35 = ws1.cell(row = i + 1, column = 106).value, \
                                  kosu_title_36 = ws1.cell(row = i + 1, column = 107).value, \
                                  kosu_division_1_36 = ws1.cell(row = i + 1, column = 108).value, \
                                  kosu_division_2_36 = ws1.cell(row = i + 1, column = 109).value, \
                                  kosu_title_37 = ws1.cell(row = i + 1, column = 110).value, \
                                  kosu_division_1_37 = ws1.cell(row = i + 1, column = 111).value, \
                                  kosu_division_2_37 = ws1.cell(row = i + 1, column = 112).value, \
                                  kosu_title_38 = ws1.cell(row = i + 1, column = 113).value, \
                                  kosu_division_1_38 = ws1.cell(row = i + 1, column = 114).value, \
                                  kosu_division_2_38 = ws1.cell(row = i + 1, column = 115).value, \
                                  kosu_title_39 = ws1.cell(row = i + 1, column = 116).value, \
                                  kosu_division_1_39 = ws1.cell(row = i + 1, column = 117).value, \
                                  kosu_division_2_39 = ws1.cell(row = i + 1, column = 118).value, \
                                  kosu_title_40 = ws1.cell(row = i + 1, column = 119).value, \
                                  kosu_division_1_40 = ws1.cell(row = i + 1, column = 120).value, \
                                  kosu_division_2_40 = ws1.cell(row = i + 1, column = 121).value, \
                                  kosu_title_41 = ws1.cell(row = i + 1, column = 122).value, \
                                  kosu_division_1_41 = ws1.cell(row = i + 1, column = 123).value, \
                                  kosu_division_2_41 = ws1.cell(row = i + 1, column = 124).value, \
                                  kosu_title_42 = ws1.cell(row = i + 1, column = 125).value, \
                                  kosu_division_1_42 = ws1.cell(row = i + 1, column = 126).value, \
                                  kosu_division_2_42 = ws1.cell(row = i + 1, column = 127).value, \
                                  kosu_title_43 = ws1.cell(row = i + 1, column = 128).value, \
                                  kosu_division_1_43 = ws1.cell(row = i + 1, column = 129).value, \
                                  kosu_division_2_43 = ws1.cell(row = i + 1, column = 130).value, \
                                  kosu_title_44 = ws1.cell(row = i + 1, column = 131).value, \
                                  kosu_division_1_44 = ws1.cell(row = i + 1, column = 132).value, \
                                  kosu_division_2_44 = ws1.cell(row = i + 1, column = 133).value, \
                                  kosu_title_45 = ws1.cell(row = i + 1, column = 134).value, \
                                  kosu_division_1_45 = ws1.cell(row = i + 1, column = 135).value, \
                                  kosu_division_2_45 = ws1.cell(row = i + 1, column = 136).value, \
                                  kosu_title_46 = ws1.cell(row = i + 1, column = 137).value, \
                                  kosu_division_1_46 = ws1.cell(row = i + 1, column = 138).value, \
                                  kosu_division_2_46 = ws1.cell(row = i + 1, column = 139).value, \
                                  kosu_title_47 = ws1.cell(row = i + 1, column = 140).value, \
                                  kosu_division_1_47 = ws1.cell(row = i + 1, column = 141).value, \
                                  kosu_division_2_47 = ws1.cell(row = i + 1, column = 142).value, \
                                  kosu_title_48 = ws1.cell(row = i + 1, column = 143).value, \
                                  kosu_division_1_48 = ws1.cell(row = i + 1, column = 144).value, \
                                  kosu_division_2_48 = ws1.cell(row = i + 1, column = 145).value, \
                                  kosu_title_49 = ws1.cell(row = i + 1, column = 146).value, \
                                  kosu_division_1_49 = ws1.cell(row = i + 1, column = 147).value, \
                                  kosu_division_2_49 = ws1.cell(row = i + 1, column = 148).value, \
                                  kosu_title_50 = ws1.cell(row = i + 1, column = 149).value, \
                                  kosu_division_1_50 = ws1.cell(row = i + 1, column = 150).value, \
                                  kosu_division_2_50 = ws1.cell(row = i + 1, column = 151).value)

        new_data1.save()


    # 工数定義区分データにレコードがない場合の処理
    else:

      # Excelからデータを読み込み
      for i in range(1, data_num):

        new_data1 = kosu_division(kosu_name = ws1.cell(row = i + 1, column = 1).value, \
                                  kosu_title_1 = ws1.cell(row = i + 1, column = 2).value, \
                                  kosu_division_1_1 = ws1.cell(row = i + 1, column = 3).value, \
                                  kosu_division_2_1 = ws1.cell(row = i + 1, column = 4).value, \
                                  kosu_title_2 = ws1.cell(row = i + 1, column = 5).value, \
                                  kosu_division_1_2 = ws1.cell(row = i + 1, column = 6).value, \
                                  kosu_division_2_2 = ws1.cell(row = i + 1, column = 7).value, \
                                  kosu_title_3 = ws1.cell(row = i + 1, column = 8).value, \
                                  kosu_division_1_3 = ws1.cell(row = i + 1, column = 9).value, \
                                  kosu_division_2_3 = ws1.cell(row = i + 1, column = 10).value, \
                                  kosu_title_4 = ws1.cell(row = i + 1, column = 11).value, \
                                  kosu_division_1_4 = ws1.cell(row = i + 1, column = 12).value, \
                                  kosu_division_2_4 = ws1.cell(row = i + 1, column = 13).value, \
                                  kosu_title_5 = ws1.cell(row = i + 1, column = 14).value, \
                                  kosu_division_1_5 = ws1.cell(row = i + 1, column = 15).value, \
                                  kosu_division_2_5 = ws1.cell(row = i + 1, column = 16).value, \
                                  kosu_title_6 = ws1.cell(row = i + 1, column = 17).value, \
                                  kosu_division_1_6 = ws1.cell(row = i + 1, column = 18).value, \
                                  kosu_division_2_6 = ws1.cell(row = i + 1, column = 19).value, \
                                  kosu_title_7 = ws1.cell(row = i + 1, column = 20).value, \
                                  kosu_division_1_7 = ws1.cell(row = i + 1, column = 21).value, \
                                  kosu_division_2_7 = ws1.cell(row = i + 1, column = 22).value, \
                                  kosu_title_8 = ws1.cell(row = i + 1, column = 23).value, \
                                  kosu_division_1_8 = ws1.cell(row = i + 1, column = 24).value, \
                                  kosu_division_2_8 = ws1.cell(row = i + 1, column = 25).value, \
                                  kosu_title_9 = ws1.cell(row = i + 1, column = 26).value, \
                                  kosu_division_1_9 = ws1.cell(row = i + 1, column = 27).value, \
                                  kosu_division_2_9 = ws1.cell(row = i + 1, column = 28).value, \
                                  kosu_title_10 = ws1.cell(row = i + 1, column = 29).value, \
                                  kosu_division_1_10 = ws1.cell(row = i + 1, column = 30).value, \
                                  kosu_division_2_10 = ws1.cell(row = i + 1, column = 31).value, \
                                  kosu_title_11 = ws1.cell(row = i + 1, column = 32).value, \
                                  kosu_division_1_11 = ws1.cell(row = i + 1, column = 33).value, \
                                  kosu_division_2_11 = ws1.cell(row = i + 1, column = 34).value, \
                                  kosu_title_12 = ws1.cell(row = i + 1, column = 35).value, \
                                  kosu_division_1_12 = ws1.cell(row = i + 1, column = 36).value, \
                                  kosu_division_2_12 = ws1.cell(row = i + 1, column = 37).value, \
                                  kosu_title_13 = ws1.cell(row = i + 1, column = 38).value, \
                                  kosu_division_1_13 = ws1.cell(row = i + 1, column = 39).value, \
                                  kosu_division_2_13 = ws1.cell(row = i + 1, column = 40).value, \
                                  kosu_title_14 = ws1.cell(row = i + 1, column = 41).value, \
                                  kosu_division_1_14 = ws1.cell(row = i + 1, column = 42).value, \
                                  kosu_division_2_14 = ws1.cell(row = i + 1, column = 43).value, \
                                  kosu_title_15 = ws1.cell(row = i + 1, column = 44).value, \
                                  kosu_division_1_15 = ws1.cell(row = i + 1, column = 45).value, \
                                  kosu_division_2_15 = ws1.cell(row = i + 1, column = 46).value, \
                                  kosu_title_16 = ws1.cell(row = i + 1, column = 47).value, \
                                  kosu_division_1_16 = ws1.cell(row = i + 1, column = 48).value, \
                                  kosu_division_2_16 = ws1.cell(row = i + 1, column = 49).value, \
                                  kosu_title_17 = ws1.cell(row = i + 1, column = 50).value, \
                                  kosu_division_1_17 = ws1.cell(row = i + 1, column = 51).value, \
                                  kosu_division_2_17 = ws1.cell(row = i + 1, column = 52).value, \
                                  kosu_title_18 = ws1.cell(row = i + 1, column = 53).value, \
                                  kosu_division_1_18 = ws1.cell(row = i + 1, column = 54).value, \
                                  kosu_division_2_18 = ws1.cell(row = i + 1, column = 55).value, \
                                  kosu_title_19 = ws1.cell(row = i + 1, column = 56).value, \
                                  kosu_division_1_19 = ws1.cell(row = i + 1, column = 57).value, \
                                  kosu_division_2_19 = ws1.cell(row = i + 1, column = 58).value, \
                                  kosu_title_20 = ws1.cell(row = i + 1, column = 59).value, \
                                  kosu_division_1_20 = ws1.cell(row = i + 1, column = 60).value, \
                                  kosu_division_2_20 = ws1.cell(row = i + 1, column = 61).value, \
                                  kosu_title_21 = ws1.cell(row = i + 1, column = 62).value, \
                                  kosu_division_1_21 = ws1.cell(row = i + 1, column = 63).value, \
                                  kosu_division_2_21 = ws1.cell(row = i + 1, column = 64).value, \
                                  kosu_title_22 = ws1.cell(row = i + 1, column = 65).value, \
                                  kosu_division_1_22 = ws1.cell(row = i + 1, column = 66).value, \
                                  kosu_division_2_22 = ws1.cell(row = i + 1, column = 67).value, \
                                  kosu_title_23 = ws1.cell(row = i + 1, column = 68).value, \
                                  kosu_division_1_23 = ws1.cell(row = i + 1, column = 69).value, \
                                  kosu_division_2_23 = ws1.cell(row = i + 1, column = 70).value, \
                                  kosu_title_24 = ws1.cell(row = i + 1, column = 71).value, \
                                  kosu_division_1_24 = ws1.cell(row = i + 1, column = 72).value, \
                                  kosu_division_2_24 = ws1.cell(row = i + 1, column = 73).value, \
                                  kosu_title_25 = ws1.cell(row = i + 1, column = 74).value, \
                                  kosu_division_1_25 = ws1.cell(row = i + 1, column = 75).value, \
                                  kosu_division_2_25 = ws1.cell(row = i + 1, column = 76).value, \
                                  kosu_title_26 = ws1.cell(row = i + 1, column = 77).value, \
                                  kosu_division_1_26 = ws1.cell(row = i + 1, column = 78).value, \
                                  kosu_division_2_26 = ws1.cell(row = i + 1, column = 79).value, \
                                  kosu_title_27 = ws1.cell(row = i + 1, column = 80).value, \
                                  kosu_division_1_27 = ws1.cell(row = i + 1, column = 81).value, \
                                  kosu_division_2_27 = ws1.cell(row = i + 1, column = 82).value, \
                                  kosu_title_28 = ws1.cell(row = i + 1, column = 83).value, \
                                  kosu_division_1_28 = ws1.cell(row = i + 1, column = 84).value, \
                                  kosu_division_2_28 = ws1.cell(row = i + 1, column = 85).value, \
                                  kosu_title_29 = ws1.cell(row = i + 1, column = 86).value, \
                                  kosu_division_1_29 = ws1.cell(row = i + 1, column = 87).value, \
                                  kosu_division_2_29 = ws1.cell(row = i + 1, column = 88).value, \
                                  kosu_title_30 = ws1.cell(row = i + 1, column = 89).value, \
                                  kosu_division_1_30 = ws1.cell(row = i + 1, column = 90).value, \
                                  kosu_division_2_30 = ws1.cell(row = i + 1, column = 91).value, \
                                  kosu_title_31 = ws1.cell(row = i + 1, column = 92).value, \
                                  kosu_division_1_31 = ws1.cell(row = i + 1, column = 93).value, \
                                  kosu_division_2_31 = ws1.cell(row = i + 1, column = 94).value, \
                                  kosu_title_32 = ws1.cell(row = i + 1, column = 95).value, \
                                  kosu_division_1_32 = ws1.cell(row = i + 1, column = 96).value, \
                                  kosu_division_2_32 = ws1.cell(row = i + 1, column = 97).value, \
                                  kosu_title_33 = ws1.cell(row = i + 1, column = 98).value, \
                                  kosu_division_1_33 = ws1.cell(row = i + 1, column = 99).value, \
                                  kosu_division_2_33 = ws1.cell(row = i + 1, column = 100).value, \
                                  kosu_title_34 = ws1.cell(row = i + 1, column = 101).value, \
                                  kosu_division_1_34 = ws1.cell(row = i + 1, column = 102).value, \
                                  kosu_division_2_34 = ws1.cell(row = i + 1, column = 103).value, \
                                  kosu_title_35 = ws1.cell(row = i + 1, column = 104).value, \
                                  kosu_division_1_35 = ws1.cell(row = i + 1, column = 105).value, \
                                  kosu_division_2_35 = ws1.cell(row = i + 1, column = 106).value, \
                                  kosu_title_36 = ws1.cell(row = i + 1, column = 107).value, \
                                  kosu_division_1_36 = ws1.cell(row = i + 1, column = 108).value, \
                                  kosu_division_2_36 = ws1.cell(row = i + 1, column = 109).value, \
                                  kosu_title_37 = ws1.cell(row = i + 1, column = 110).value, \
                                  kosu_division_1_37 = ws1.cell(row = i + 1, column = 111).value, \
                                  kosu_division_2_37 = ws1.cell(row = i + 1, column = 112).value, \
                                  kosu_title_38 = ws1.cell(row = i + 1, column = 113).value, \
                                  kosu_division_1_38 = ws1.cell(row = i + 1, column = 114).value, \
                                  kosu_division_2_38 = ws1.cell(row = i + 1, column = 115).value, \
                                  kosu_title_39 = ws1.cell(row = i + 1, column = 116).value, \
                                  kosu_division_1_39 = ws1.cell(row = i + 1, column = 117).value, \
                                  kosu_division_2_39 = ws1.cell(row = i + 1, column = 118).value, \
                                  kosu_title_40 = ws1.cell(row = i + 1, column = 119).value, \
                                  kosu_division_1_40 = ws1.cell(row = i + 1, column = 120).value, \
                                  kosu_division_2_40 = ws1.cell(row = i + 1, column = 121).value, \
                                  kosu_title_41 = ws1.cell(row = i + 1, column = 122).value, \
                                  kosu_division_1_41 = ws1.cell(row = i + 1, column = 123).value, \
                                  kosu_division_2_41 = ws1.cell(row = i + 1, column = 124).value, \
                                  kosu_title_42 = ws1.cell(row = i + 1, column = 125).value, \
                                  kosu_division_1_42 = ws1.cell(row = i + 1, column = 126).value, \
                                  kosu_division_2_42 = ws1.cell(row = i + 1, column = 127).value, \
                                  kosu_title_43 = ws1.cell(row = i + 1, column = 128).value, \
                                  kosu_division_1_43 = ws1.cell(row = i + 1, column = 129).value, \
                                  kosu_division_2_43 = ws1.cell(row = i + 1, column = 130).value, \
                                  kosu_title_44 = ws1.cell(row = i + 1, column = 131).value, \
                                  kosu_division_1_44 = ws1.cell(row = i + 1, column = 132).value, \
                                  kosu_division_2_44 = ws1.cell(row = i + 1, column = 133).value, \
                                  kosu_title_45 = ws1.cell(row = i + 1, column = 134).value, \
                                  kosu_division_1_45 = ws1.cell(row = i + 1, column = 135).value, \
                                  kosu_division_2_45 = ws1.cell(row = i + 1, column = 136).value, \
                                  kosu_title_46 = ws1.cell(row = i + 1, column = 137).value, \
                                  kosu_division_1_46 = ws1.cell(row = i + 1, column = 138).value, \
                                  kosu_division_2_46 = ws1.cell(row = i + 1, column = 139).value, \
                                  kosu_title_47 = ws1.cell(row = i + 1, column = 140).value, \
                                  kosu_division_1_47 = ws1.cell(row = i + 1, column = 141).value, \
                                  kosu_division_2_47 = ws1.cell(row = i + 1, column = 142).value, \
                                  kosu_title_48 = ws1.cell(row = i + 1, column = 143).value, \
                                  kosu_division_1_48 = ws1.cell(row = i + 1, column = 144).value, \
                                  kosu_division_2_48 = ws1.cell(row = i + 1, column = 145).value, \
                                  kosu_title_49 = ws1.cell(row = i + 1, column = 146).value, \
                                  kosu_division_1_49 = ws1.cell(row = i + 1, column = 147).value, \
                                  kosu_division_2_49 = ws1.cell(row = i + 1, column = 148).value, \
                                  kosu_title_50 = ws1.cell(row = i + 1, column = 149).value, \
                                  kosu_division_1_50 = ws1.cell(row = i + 1, column = 150).value, \
                                  kosu_division_2_50 = ws1.cell(row = i + 1, column = 151).value)

        new_data1.save()


    # 一時ファイル削除
    os.remove('def_file_path.xlsx')



    # 工数定義区分ファイル定義
    uploaded_file = request.FILES['setting_file']

    # 一時的なファイルをサーバー上に作成
    with open('setting_file_path.xlsx', 'wb+') as destination:

      # アップロードしたファイルを一時ファイルに書き込み
      for chunk in uploaded_file.chunks():
        destination.write(chunk)

    # 一時ファイルを開く
    wb2 = openpyxl.load_workbook('setting_file_path.xlsx')

    # 一番最初のシートを指定
    ws2 = wb2.worksheets[0]

    # レコード数取得
    data_num = ws2.max_row


    # 管理者設定データにレコードがある場合の処理
    if administrator_data.objects.exists():

      # 管理者設定データ取得
      setting_obj_get = administrator_data.objects.all()
      # 取得した管理者設定データを消す
      setting_obj_get.delete()


    # Excelからデータを読み込むループ
    for i in range(1, data_num):

      # Excelからデータを読み込み
      new_data2 = administrator_data(menu_row = ws2.cell(row = i + 1, column = 1).value)

      new_data2.save()


    # 一時ファイル削除
    os.remove('setting_file_path.xlsx')


    # このページをリダイレクト
    return redirect(to = '/login')
  


  # HTMLに渡す辞書
  context = {
    'title' : 'ヘルプ',
    'form' : form,
    'display' : display,
    }
  


  # 指定したHTMLに辞書を渡して表示を完成させる
  return render(request, 'kosu/help.html', context)





#--------------------------------------------------------------------------------------------------------





# ストップウォッチ画面定義
def stopwatch(request):

  # セッションにログインした従業員番号がない場合の処理
  if not request.session.get('login_No'):
    # 未ログインならログインページへ飛ぶ
    return redirect('/login')
  
  try:
    # ログイン者の情報取得
    member_data = member.objects.get(employee_no=request.session['login_No'])
  # セッション値から人員情報取得できない場合の処理
  except member.DoesNotExist:
    # セッション削除
    request.session.clear()
    # ログインページに戻る
    return redirect('/login')
   

   
  # HTMLに渡す辞書
  context = {
    'title' : 'ストップウォッチ',
    }
  
  # 指定したHTMLに辞書を渡して表示を完成させる
  return render(request, 'kosu/stopwatch.html', context)





#--------------------------------------------------------------------------------------------------------





# 半角文字入力チェック関数
def has_non_halfwidth_characters(input_string):

  for char in input_string:

    # Unicodeの文字幅カテゴリ 'Na' は「ナロー（半角）」を意味します。
    if unicodedata.east_asian_width(char) != 'Na':
      # 全角文字かそれ以外（例: 全角スペース、絵文字など）が検出されたらTrueを返す
      return True

    return False





#--------------------------------------------------------------------------------------------------------









